from __future__ import annotations

import argparse
import base64
import binascii
import csv
import difflib
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import socket
import sqlite3
import tempfile
import traceback
import zipfile
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    Workbook = None
    load_workbook = None
    Font = None
    PatternFill = None
    InvalidFileException = ValueError

try:
    from PIL import Image, ImageOps, UnidentifiedImageError
    Image.MAX_IMAGE_PIXELS = 12_000_000
except ImportError:
    Image = None
    ImageOps = None
    UnidentifiedImageError = ValueError


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("PPWORK_DATA_DIR") or BASE_DIR).expanduser().resolve()
STATIC_DIR = BASE_DIR / "static"
ASSETS_DIR = STATIC_DIR / "assets"
BACKUP_DIR = DATA_DIR / "backups"
LOG_DIR = DATA_DIR / "logs"
LOG_FILE = LOG_DIR / "app.log"
APP_VERSION = "0.14.3"
SCHEMA_VERSION = "2026-07-07-0.13.0"
DEFAULT_DEPARTMENT = "parts"
DEPARTMENTS = {
    "parts": {"label": "Parts", "db": DATA_DIR / "parts.db", "seed": True},
    "service": {"label": "Service", "db": DATA_DIR / "service.db", "seed": False},
}
ADMIN_PASSWORD_HASH = "1fe8359332f00ab7dde21a97ba3603eb06b8f68266c0a7e9e7582f0efc039dd0"
SEED_PATH = BASE_DIR / "seed-data.json"
EMPTY_INSTALL_MARKER = DATA_DIR / ".counterflow-empty-install"
MAX_JSON_BYTES = 8 * 1024 * 1024
MAX_LOGO_BYTES = 4 * 1024 * 1024
MAX_LOGO_DIMENSION = 512
MIN_LOGO_DIMENSION = 16
LOGO_TYPES = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/gif": ".gif",
}
PART_EXPORT_COLUMNS = [
    "id", "brand", "family", "model", "category", "yearStart", "yearEnd",
    "make", "fitmentModel", "unitType", "reviewStatus", "reviewNote", "item",
    "buttonText", "partNumber", "oldPartNumber", "newPartNumber", "alternateNumbers",
    "aftermarketNumbers", "vendor", "tags", "fitmentNotes", "attachmentUrl", "notes",
    "source", "sortOrder",
]


ROLE_ORDER = ["counter", "manager", "admin"]
PERMISSION_DEFINITIONS = {
    "import": "Import parts",
    "export": "Export parts",
    "brandEdit": "Brand editing",
    "employeeEdit": "Employee editing",
    "permanentBrandDelete": "Permanent saved-brand removal",
}
DEFAULT_ROLE_PERMISSIONS = {
    "counter": [],
    "manager": ["import", "export", "brandEdit"],
    "admin": list(PERMISSION_DEFINITIONS),
}

DEFAULT_SETTINGS = {

    "dealershipName": "Independence County Offroad",
    "locationName": "",
    "partsDepartmentLabel": "Parts",
    "serviceDepartmentLabel": "Service",
    "rolePermissions": json.dumps(DEFAULT_ROLE_PERMISSIONS, sort_keys=True),
}

SERVICE_RESOURCE_TYPES = {
    "labor_template",
    "favorite_kit",
    "model_note",
    "seasonal_package",
}

EMPLOYEE_ROLES = {
    "counter",
    "manager",
    "admin",
}


def normalize_department(value: object) -> str:
    department = str(value or "").strip().lower()
    if department in DEPARTMENTS:
        return department
    return DEFAULT_DEPARTMENT


def db_path_for_department(department: str) -> Path:
    return DEPARTMENTS[normalize_department(department)]["db"]


def department_label(department: str) -> str:
    return DEPARTMENTS[normalize_department(department)]["label"]


def valid_admin_password(value: object) -> bool:
    digest = hashlib.sha256(clean_text(value).encode("utf-8")).hexdigest()
    return secrets.compare_digest(digest, ADMIN_PASSWORD_HASH)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def log_event(level: str, message: str, **context: object) -> None:
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        payload = {
            "time": now_iso(),
            "level": level,
            "message": message,
            **{key: value for key, value in context.items() if value not in (None, "")},
        }
        with LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(payload, default=str) + "\n")
    except OSError:
        pass


def log_exception(error: BaseException, **context: object) -> None:
    log_event(
        "error",
        str(error) or error.__class__.__name__,
        errorType=error.__class__.__name__,
        traceback="".join(traceback.format_exception(error)).strip(),
        **context,
    )


def friendly_error_message(error: BaseException) -> tuple[HTTPStatus, str]:
    if isinstance(error, sqlite3.OperationalError):
        message = str(error).lower()
        if "locked" in message or "busy" in message:
            return HTTPStatus.SERVICE_UNAVAILABLE, "The database is busy. Wait a few seconds and try again."
        if "readonly" in message or "permission" in message:
            return HTTPStatus.SERVICE_UNAVAILABLE, "The database could not be written. Check that the app folder is not read-only."
    if isinstance(error, FileNotFoundError):
        return HTTPStatus.NOT_FOUND, "The requested file could not be found."
    return HTTPStatus.INTERNAL_SERVER_ERROR, "Something went wrong. The error was logged for review."


def connect(department: str = DEFAULT_DEPARTMENT) -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path_for_department(department))
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def init_db() -> None:
    for department, config in DEPARTMENTS.items():
        with connect(department) as db:
            db.executescript(
                """
                CREATE TABLE IF NOT EXISTS brands (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    accent TEXT NOT NULL DEFAULT '#2563eb',
                    logo TEXT NOT NULL DEFAULT '',
                    category TEXT NOT NULL DEFAULT '',
                    default_family TEXT NOT NULL DEFAULT '',
                    default_model TEXT NOT NULL DEFAULT '',
                    default_category TEXT NOT NULL DEFAULT '',
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    active INTEGER NOT NULL DEFAULT 1,
                    archived_at TEXT NOT NULL DEFAULT '',
                    archive_note TEXT NOT NULL DEFAULT '',
                    deleted_at TEXT NOT NULL DEFAULT '',
                    deleted_name TEXT NOT NULL DEFAULT ''
                );

                CREATE TABLE IF NOT EXISTS parts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
                    family TEXT NOT NULL DEFAULT '',
                    model TEXT NOT NULL DEFAULT '',
                    category TEXT NOT NULL DEFAULT '',
                    year_start INTEGER NOT NULL DEFAULT 0,
                    year_end INTEGER NOT NULL DEFAULT 0,
                    make TEXT NOT NULL DEFAULT '',
                    fitment_model TEXT NOT NULL DEFAULT '',
                    unit_type TEXT NOT NULL DEFAULT '',
                    review_status TEXT NOT NULL DEFAULT 'approved',
                    review_note TEXT NOT NULL DEFAULT '',
                    item TEXT NOT NULL,
                    button_text TEXT NOT NULL DEFAULT '',
                    part_number TEXT NOT NULL DEFAULT '',
                    old_part_number TEXT NOT NULL DEFAULT '',
                    new_part_number TEXT NOT NULL DEFAULT '',
                    alternate_numbers TEXT NOT NULL DEFAULT '',
                    aftermarket_numbers TEXT NOT NULL DEFAULT '',
                    vendor TEXT NOT NULL DEFAULT '',
                    tags TEXT NOT NULL DEFAULT '',
                    fitment_notes TEXT NOT NULL DEFAULT '',
                    attachment_url TEXT NOT NULL DEFAULT '',
                    notes TEXT NOT NULL DEFAULT '',
                    source TEXT NOT NULL DEFAULT '',
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE INDEX IF NOT EXISTS idx_parts_brand ON parts(brand_id);
                CREATE INDEX IF NOT EXISTS idx_parts_lookup ON parts(active, family, model, category, item);

                CREATE TABLE IF NOT EXISTS part_audit (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    part_id INTEGER NOT NULL DEFAULT 0,
                    action TEXT NOT NULL,
                    summary TEXT NOT NULL DEFAULT '',
                    changes_json TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_part_audit_part ON part_audit(part_id, created_at);
                CREATE TABLE IF NOT EXISTS service_resources (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    resource_type TEXT NOT NULL,
                    title TEXT NOT NULL,
                    brand TEXT NOT NULL DEFAULT '',
                    model TEXT NOT NULL DEFAULT '',
                    unit_type TEXT NOT NULL DEFAULT '',
                    season TEXT NOT NULL DEFAULT '',
                    content TEXT NOT NULL DEFAULT '',
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_service_resources_type ON service_resources(active, resource_type, sort_order);

                CREATE TABLE IF NOT EXISTS employees (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    username TEXT NOT NULL DEFAULT '',
                    role TEXT NOT NULL DEFAULT 'counter',
                    pin_hash TEXT NOT NULL DEFAULT '',
                    password_hash TEXT NOT NULL DEFAULT '',
                    allowed_departments TEXT NOT NULL DEFAULT 'parts,service',
                    location_scope TEXT NOT NULL DEFAULT '',
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_active_name ON employees(name COLLATE NOCASE) WHERE active = 1;

                CREATE TABLE IF NOT EXISTS employee_favorites (
                    employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                    department TEXT NOT NULL,
                    part_id INTEGER NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY (employee_id, department, part_id)
                );
                CREATE INDEX IF NOT EXISTS idx_employee_favorites_department ON employee_favorites(department, part_id);

                CREATE TABLE IF NOT EXISTS employee_copy_activity (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL DEFAULT 0,
                    employee_name TEXT NOT NULL DEFAULT '',
                    department TEXT NOT NULL,
                    part_id INTEGER NOT NULL DEFAULT 0,
                    brand TEXT NOT NULL DEFAULT '',
                    item TEXT NOT NULL DEFAULT '',
                    part_number TEXT NOT NULL DEFAULT '',
                    copied_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_employee_copy_activity_department ON employee_copy_activity(department, copied_at);
                CREATE INDEX IF NOT EXISTS idx_employee_copy_activity_employee ON employee_copy_activity(employee_id, copied_at);

                CREATE TABLE IF NOT EXISTS employee_sessions (
                    token TEXT PRIMARY KEY,
                    employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                    created_at TEXT NOT NULL,
                    last_seen_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_employee_sessions_employee ON employee_sessions(employee_id, last_seen_at);

                CREATE TABLE IF NOT EXISTS schema_migrations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    app_version TEXT NOT NULL DEFAULT '',
                    department TEXT NOT NULL DEFAULT '',
                    applied_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS app_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL DEFAULT '',
                    updated_at TEXT NOT NULL
                );
                """
            )
            ensure_brand_archive_columns(db)
            ensure_part_reference_columns(db)
            ensure_part_audit_table(db)
            ensure_employee_access_columns(db)
            record_schema_migration(db, department, "baseline-schema")
            record_schema_migration(db, department, "employee-sessions-and-maintenance")
            record_schema_migration(db, department, "shared-data-folder-support")
            record_schema_migration(db, department, "deployment-packaging-and-autostart")
            record_schema_migration(db, department, "settings-tabs-admin-access")
            record_schema_migration(db, department, "counter-speed-and-maintenance-tools")
            record_schema_migration(db, department, "first-run-admin-and-role-permissions")
            db.execute("CREATE INDEX IF NOT EXISTS idx_brands_active ON brands(active, sort_order)")
            brand_count = db.execute("SELECT COUNT(*) FROM brands").fetchone()[0]
            part_count = db.execute("SELECT COUNT(*) FROM parts").fetchone()[0]

            if config["seed"] and brand_count == 0 and part_count == 0 and SEED_PATH.exists() and not EMPTY_INSTALL_MARKER.exists():
                seed_database(db)


def seed_database(db: sqlite3.Connection) -> None:
    data = json.loads(SEED_PATH.read_text(encoding="utf-8"))
    stamp = now_iso()

    for index, brand in enumerate(data.get("brands", []), start=1):
        db.execute(
            """
            INSERT INTO brands (name, accent, logo, sort_order)
            VALUES (?, ?, ?, ?)
            """,
            (
                clean_text(brand.get("name")),
                clean_text(brand.get("accent")) or "#2563eb",
                clean_text(brand.get("logo")),
                index,
            ),
        )

    brand_ids = {row["name"]: row["id"] for row in db.execute("SELECT id, name FROM brands")}
    for part in data.get("parts", []):
        brand_name = clean_text(part.get("brand"))
        brand_id = brand_ids.get(brand_name)
        if not brand_id:
            continue

        db.execute(
            """
            INSERT INTO parts (
                brand_id, family, model, category, item, button_text, part_number,
                notes, source, sort_order, active, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                brand_id,
                clean_text(part.get("family")),
                clean_text(part.get("model")),
                clean_text(part.get("category")),
                clean_text(part.get("item")) or "Untitled Part",
                clean_text(part.get("buttonText")),
                clean_text(part.get("partNumber")),
                clean_text(part.get("notes")),
                clean_text(part.get("source")),
                int(part.get("sortOrder") or 0),
                1 if part.get("active", True) else 0,
                stamp,
                stamp,
            ),
        )


def record_schema_migration(db: sqlite3.Connection, department: str, name: str) -> None:
    db.execute(
        """
        INSERT OR IGNORE INTO schema_migrations (name, app_version, department, applied_at)
        VALUES (?, ?, ?, ?)
        """,
        (f"{department}:{name}", APP_VERSION, department, now_iso()),
    )


def ensure_brand_archive_columns(db: sqlite3.Connection) -> None:
    columns = {row["name"] for row in db.execute("PRAGMA table_info(brands)").fetchall()}
    if "active" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN active INTEGER NOT NULL DEFAULT 1")
    if "archived_at" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN archived_at TEXT NOT NULL DEFAULT ''")
    if "deleted_at" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN deleted_at TEXT NOT NULL DEFAULT ''")
    if "deleted_name" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN deleted_name TEXT NOT NULL DEFAULT ''")
    if "category" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN category TEXT NOT NULL DEFAULT ''")
    if "default_family" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN default_family TEXT NOT NULL DEFAULT ''")
    if "default_model" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN default_model TEXT NOT NULL DEFAULT ''")
    if "default_category" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN default_category TEXT NOT NULL DEFAULT ''")
    if "archive_note" not in columns:
        db.execute("ALTER TABLE brands ADD COLUMN archive_note TEXT NOT NULL DEFAULT ''")



def ensure_part_reference_columns(db: sqlite3.Connection) -> None:
    columns = {row["name"] for row in db.execute("PRAGMA table_info(parts)").fetchall()}
    reference_columns = {
        "year_start": "INTEGER NOT NULL DEFAULT 0",
        "year_end": "INTEGER NOT NULL DEFAULT 0",
        "make": "TEXT NOT NULL DEFAULT ''",
        "fitment_model": "TEXT NOT NULL DEFAULT ''",
        "unit_type": "TEXT NOT NULL DEFAULT ''",
        "review_status": "TEXT NOT NULL DEFAULT 'approved'",
        "review_note": "TEXT NOT NULL DEFAULT ''",
        "old_part_number": "TEXT NOT NULL DEFAULT ''",
        "new_part_number": "TEXT NOT NULL DEFAULT ''",
        "alternate_numbers": "TEXT NOT NULL DEFAULT ''",
        "aftermarket_numbers": "TEXT NOT NULL DEFAULT ''",
        "vendor": "TEXT NOT NULL DEFAULT ''",
        "tags": "TEXT NOT NULL DEFAULT ''",
        "fitment_notes": "TEXT NOT NULL DEFAULT ''",
        "attachment_url": "TEXT NOT NULL DEFAULT ''",
    }
    for name, definition in reference_columns.items():
        if name not in columns:
            db.execute(f"ALTER TABLE parts ADD COLUMN {name} {definition}")


def ensure_part_audit_table(db: sqlite3.Connection) -> None:
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS part_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL DEFAULT 0,
            action TEXT NOT NULL,
            summary TEXT NOT NULL DEFAULT '',
            changes_json TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_part_audit_part ON part_audit(part_id, created_at);
        """
    )

def ensure_employee_access_columns(db: sqlite3.Connection) -> None:
    columns = {row["name"] for row in db.execute("PRAGMA table_info(employees)").fetchall()}
    additions = {
        "username": "TEXT NOT NULL DEFAULT ''",
        "password_hash": "TEXT NOT NULL DEFAULT ''",
        "allowed_departments": "TEXT NOT NULL DEFAULT 'parts,service'",
        "location_scope": "TEXT NOT NULL DEFAULT ''",
    }
    for name, definition in additions.items():
        if name not in columns:
            db.execute(f"ALTER TABLE employees ADD COLUMN {name} {definition}")
    db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_active_username ON employees(username COLLATE NOCASE) WHERE active = 1 AND username <> ''")

def backup_database(department: str, reason: str) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    normalized = normalize_department(department)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_reason = re.sub(r"[^a-z0-9-]+", "-", reason.lower()).strip("-") or "backup"
    target = BACKUP_DIR / f"{normalized}-{safe_reason}-{stamp}.db"
    shutil.copy2(db_path_for_department(normalized), target)
    return target


def create_daily_startup_backups() -> None:
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    for department in DEPARTMENTS:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        if any(BACKUP_DIR.glob(f"{department}-scheduled-daily-{today}-*.db")):
            continue
        try:
            backup_database(department, f"scheduled-daily-{today}")
        except OSError as error:
            print(f"Could not create scheduled backup for {department}: {error}")


def backup_path_from_name(department: str, filename: str) -> Path | None:
    name = Path(clean_text(filename)).name
    normalized = normalize_department(department)
    if not name.startswith(f"{normalized}-") or not name.endswith(".db"):
        return None
    target = (BACKUP_DIR / name).resolve()
    backup_root = BACKUP_DIR.resolve()
    if not str(target).startswith(str(backup_root)) or not target.exists() or not target.is_file():
        return None
    return target


def csv_value(row: dict[str, str], *keys: str) -> str:
    lower = {key.lower(): value for key, value in row.items()}
    for key in keys:
        if key in row:
            return clean_text(row.get(key))
        if key.lower() in lower:
            return clean_text(lower.get(key.lower()))
    return ""


def parse_int(value: object, default: int = 0) -> int:
    try:
        return int(clean_text(value) or default)
    except ValueError:
        return default


def part_payload_from_csv(row: dict[str, str]) -> dict[str, str | int]:
    return {
        "brand": csv_value(row, "brand"),
        "family": csv_value(row, "family"),
        "model": csv_value(row, "model"),
        "category": csv_value(row, "category"),
        "yearStart": parse_int(csv_value(row, "yearStart", "year_start", "yearFrom", "year_from"), 0),
        "yearEnd": parse_int(csv_value(row, "yearEnd", "year_end", "yearTo", "year_to"), 0),
        "make": csv_value(row, "make"),
        "fitmentModel": csv_value(row, "fitmentModel", "fitment_model", "unitModel", "unit_model"),
        "unitType": csv_value(row, "unitType", "unit_type", "vehicleType", "vehicle_type"),
        "reviewStatus": csv_value(row, "reviewStatus", "review_status") or "approved",
        "reviewNote": csv_value(row, "reviewNote", "review_note"),
        "item": csv_value(row, "item") or "Untitled Part",
        "buttonText": csv_value(row, "buttonText", "button_text"),
        "partNumber": csv_value(row, "partNumber", "part_number"),
        "oldPartNumber": csv_value(row, "oldPartNumber", "old_part_number"),
        "newPartNumber": csv_value(row, "newPartNumber", "new_part_number"),
        "alternateNumbers": csv_value(row, "alternateNumbers", "alternate_numbers"),
        "aftermarketNumbers": csv_value(row, "aftermarketNumbers", "aftermarket_numbers"),
        "vendor": csv_value(row, "vendor"),
        "tags": csv_value(row, "tags"),
        "fitmentNotes": csv_value(row, "fitmentNotes", "fitment_notes"),
        "attachmentUrl": csv_value(row, "attachmentUrl", "attachment_url"),
        "notes": csv_value(row, "notes"),
        "source": csv_value(row, "source"),
        "sortOrder": parse_int(csv_value(row, "sortOrder", "sort_order"), 0),
    }


def rows_from_worksheet(workbook) -> list[dict[str, str]]:
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file needs a header row.")
    headers = [clean_text(value) for value in rows[0]]
    if not any(headers):
        raise ValueError("Excel file needs a header row.")
    records: list[dict[str, str]] = []
    for values in rows[1:]:
        if not values or not any(clean_text(value) for value in values):
            continue
        record = {header: clean_text(value) for header, value in zip(headers, values) if header}
        records.append(record)
    return records


def search_values_for_row(row: sqlite3.Row) -> list[str]:
    year_start = int(row["year_start"] or 0)
    year_end = int(row["year_end"] or 0)
    year_label = ""
    if year_start and year_end:
        year_label = f"{year_start}-{year_end}" if year_start != year_end else str(year_start)
    elif year_start:
        year_label = f"{year_start}+"
    elif year_end:
        year_label = f"through {year_end}"
    return [
        row["brand"], row["family"], row["model"], row["category"], year_label,
        row["make"], row["fitment_model"], row["unit_type"], row["item"], row["button_text"],
        row["part_number"], row["old_part_number"], row["new_part_number"], row["alternate_numbers"],
        row["aftermarket_numbers"], row["vendor"], row["tags"], row["fitment_notes"],
        row["attachment_url"], row["notes"], row["source"],
    ]


def normalized_search(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def token_matches_candidate(token: str, candidate: str) -> bool:
    if not token:
        return True
    if token in candidate or candidate in token:
        return True
    if len(token) < 3 or len(candidate) < 3:
        return False
    ratio = difflib.SequenceMatcher(None, token, candidate).ratio()
    return ratio >= 0.74


def row_matches_search(row: sqlite3.Row, search: str) -> bool:
    query = clean_text(search)
    if not query:
        return True
    values = search_values_for_row(row)
    combined = normalized_search(" ".join(clean_text(value) for value in values))
    normalized_query = normalized_search(query)
    if normalized_query and normalized_query in combined:
        return True
    tokens = [normalized_search(token) for token in re.split(r"\s+", query) if normalized_search(token)]
    candidates = []
    for value in values:
        candidates.extend(normalized_search(token) for token in re.split(r"[^A-Za-z0-9]+", clean_text(value)) if normalized_search(token))
    return all(any(token_matches_candidate(token, candidate) for candidate in candidates) for token in tokens)


def record_part_audit(db: sqlite3.Connection, part_id: int, action: str, summary: str = "", before: dict | None = None, after: dict | None = None) -> None:
    db.execute(
        """
        INSERT INTO part_audit (part_id, action, summary, changes_json, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (part_id, action, summary, json.dumps({"before": before or {}, "after": after or {}}, sort_keys=True), now_iso()),
    )


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_to_part(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "brandId": row["brand_id"],
        "brand": row["brand"],
        "accent": row["accent"],
        "logo": row["logo"],
        "family": row["family"],
        "model": row["model"],
        "category": row["category"],
        "yearStart": row["year_start"],
        "yearEnd": row["year_end"],
        "make": row["make"],
        "fitmentModel": row["fitment_model"],
        "unitType": row["unit_type"],
        "reviewStatus": row["review_status"],
        "reviewNote": row["review_note"],
        "item": row["item"],
        "buttonText": row["button_text"],
        "partNumber": row["part_number"],
        "oldPartNumber": row["old_part_number"],
        "newPartNumber": row["new_part_number"],
        "alternateNumbers": row["alternate_numbers"],
        "aftermarketNumbers": row["aftermarket_numbers"],
        "vendor": row["vendor"],
        "tags": row["tags"],
        "fitmentNotes": row["fitment_notes"],
        "attachmentUrl": row["attachment_url"],
        "notes": row["notes"],
        "source": row["source"],
        "sortOrder": row["sort_order"],
        "active": bool(row["active"]),
        "updatedAt": row["updated_at"],
    }



def part_export_row(row: sqlite3.Row) -> dict[str, object]:
    return {
        "id": row["id"],
        "brand": row["brand"],
        "family": row["family"],
        "model": row["model"],
        "category": row["category"],
        "yearStart": row["year_start"],
        "yearEnd": row["year_end"],
        "make": row["make"],
        "fitmentModel": row["fitment_model"],
        "unitType": row["unit_type"],
        "reviewStatus": row["review_status"],
        "reviewNote": row["review_note"],
        "item": row["item"],
        "buttonText": row["button_text"],
        "partNumber": row["part_number"],
        "oldPartNumber": row["old_part_number"],
        "newPartNumber": row["new_part_number"],
        "alternateNumbers": row["alternate_numbers"],
        "aftermarketNumbers": row["aftermarket_numbers"],
        "vendor": row["vendor"],
        "tags": row["tags"],
        "fitmentNotes": row["fitment_notes"],
        "attachmentUrl": row["attachment_url"],
        "notes": row["notes"],
        "source": row["source"],
        "sortOrder": row["sort_order"],
    }


def prepare_logo_image(image_bytes: bytes) -> tuple[bytes, str, int, int, bool]:
    if Image is None or ImageOps is None:
        return image_bytes, "", 0, 0, False
    try:
        with Image.open(io.BytesIO(image_bytes)) as original:
            image = ImageOps.exif_transpose(original)
            width, height = image.size
            if width < MIN_LOGO_DIMENSION or height < MIN_LOGO_DIMENSION:
                raise ValueError("Logo image is too small to use.")
            if width * height > 12_000_000:
                raise ValueError("Logo image dimensions are too large.")
            image.thumbnail((MAX_LOGO_DIMENSION, MAX_LOGO_DIMENSION), Image.Resampling.LANCZOS)
            if image.mode not in {"RGB", "RGBA"}:
                image = image.convert("RGBA")
            output = io.BytesIO()
            image.save(output, format="PNG", optimize=True)
            resized = output.getvalue()
            if not resized:
                raise ValueError("Logo file could not be processed.")
            return resized, ".png", image.width, image.height, (width, height) != image.size
    except UnidentifiedImageError as error:
        raise ValueError("Logo upload must be a readable image file.") from error
    except OSError as error:
        raise ValueError("Logo upload could not be processed.") from error


def row_to_service_resource(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "type": row["resource_type"],
        "title": row["title"],
        "brand": row["brand"],
        "model": row["model"],
        "unitType": row["unit_type"],
        "season": row["season"],
        "content": row["content"],
        "sortOrder": row["sort_order"],
        "active": bool(row["active"]),
        "updatedAt": row["updated_at"],
    }


def normalize_review_status(value: object) -> str:
    status = clean_text(value).lower().replace("_", "-")
    return "needs-review" if status in {"needs-review", "review", "pending"} else "approved"


def normalize_service_resource_type(value: object) -> str:
    resource_type = clean_text(value).lower().replace("-", "_")
    if resource_type not in SERVICE_RESOURCE_TYPES:
        raise ValueError("Choose a valid service workflow type.")
    return resource_type


def settings_connection() -> sqlite3.Connection:
    return connect(DEFAULT_DEPARTMENT)


def read_app_settings() -> dict[str, str]:
    with settings_connection() as db:
        rows = db.execute("SELECT key, value FROM app_settings").fetchall()
    settings = dict(DEFAULT_SETTINGS)
    settings.update({row["key"]: row["value"] for row in rows})
    return settings


def normalize_role_permissions(value: object) -> dict[str, list[str]]:
    if isinstance(value, str):
        try:
            raw = json.loads(value) if value else {}
        except json.JSONDecodeError:
            raw = {}
    elif isinstance(value, dict):
        raw = value
    else:
        raw = {}

    normalized: dict[str, list[str]] = {}
    for role in ROLE_ORDER:
        values = raw.get(role, DEFAULT_ROLE_PERMISSIONS.get(role, [])) if isinstance(raw, dict) else []
        if not isinstance(values, list):
            values = []
        permissions = []
        for permission in values:
            key = clean_text(permission)
            if key in PERMISSION_DEFINITIONS and key not in permissions:
                permissions.append(key)
        if role == "admin":
            permissions = list(PERMISSION_DEFINITIONS)
        normalized[role] = permissions
    return normalized


def role_permissions_json(value: object) -> str:
    return json.dumps(normalize_role_permissions(value), sort_keys=True)


def read_role_permissions() -> dict[str, list[str]]:
    return normalize_role_permissions(read_app_settings().get("rolePermissions"))


def permission_actions_payload() -> list[dict[str, str]]:
    return [
        {"key": key, "label": label}
        for key, label in PERMISSION_DEFINITIONS.items()
    ]


def app_settings_payload() -> dict[str, object]:
    settings: dict[str, object] = dict(read_app_settings())
    settings["rolePermissions"] = normalize_role_permissions(settings.get("rolePermissions"))
    settings["permissionActions"] = permission_actions_payload()
    settings["employeeRoles"] = ROLE_ORDER
    return settings


def admin_employee_exists() -> bool:
    with settings_connection() as db:
        row = db.execute("SELECT COUNT(*) FROM employees WHERE active = 1 AND role = 'admin'").fetchone()
    return bool(row and row[0])


def normalize_employee_role(value: object) -> str:
    role = clean_text(value).lower()
    return role if role in EMPLOYEE_ROLES else "counter"


def hash_password(value: object) -> str:
    password = clean_text(value)
    if not password:
        return ""
    salt = secrets.token_hex(8)
    digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return f"{salt}:{digest}"


def verify_password(value: object, stored_hash: str) -> bool:
    stored = clean_text(stored_hash)
    password = clean_text(value)
    if not stored or not password or ":" not in stored:
        return False
    salt, expected = stored.split(":", 1)
    digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return secrets.compare_digest(digest, expected)


def normalize_allowed_departments(value: object) -> str:
    if isinstance(value, list):
        raw_values = value
    else:
        raw_values = str(value or "parts,service").split(",")
    allowed: list[str] = []
    for item in raw_values:
        department = clean_text(item).lower()
        if department in DEPARTMENTS and department not in allowed:
            allowed.append(department)
    return ",".join(allowed or [DEFAULT_DEPARTMENT])


def employee_allowed_departments(row: sqlite3.Row) -> list[str]:
    if row["role"] in {"manager", "admin"}:
        return list(DEPARTMENTS.keys())
    raw = row["allowed_departments"] if "allowed_departments" in row.keys() else "parts,service"
    departments: list[str] = []
    for item in clean_text(raw).split(","):
        department = clean_text(item).lower()
        if department in DEPARTMENTS and department not in departments:
            departments.append(department)
    return departments or [DEFAULT_DEPARTMENT]
def hash_pin(value: object) -> str:
    pin = clean_text(value)
    if not pin:
        return ""
    salt = secrets.token_hex(8)
    digest = hashlib.sha256(f"{salt}:{pin}".encode("utf-8")).hexdigest()
    return f"{salt}:{digest}"


def verify_pin(value: object, stored_hash: str) -> bool:
    stored = clean_text(stored_hash)
    if not stored:
        return True
    pin = clean_text(value)
    if not pin or ":" not in stored:
        return False
    salt, expected = stored.split(":", 1)
    digest = hashlib.sha256(f"{salt}:{pin}".encode("utf-8")).hexdigest()
    return secrets.compare_digest(digest, expected)


def row_to_employee(row: sqlite3.Row, session_token: str = "") -> dict[str, object]:
    payload = {
        "id": row["id"],
        "name": row["name"],
        "username": row["username"] if "username" in row.keys() else "",
        "role": row["role"],
        "active": bool(row["active"]),
        "hasPin": bool(row["pin_hash"]),
        "hasPassword": bool(row["password_hash"] if "password_hash" in row.keys() else ""),
        "allowedDepartments": employee_allowed_departments(row),
        "locationScope": row["location_scope"] if "location_scope" in row.keys() else "",
        "updatedAt": row["updated_at"],
    }
    if session_token:
        payload["sessionToken"] = session_token
    return payload


def create_employee_session(employee_id: int) -> str:
    token = secrets.token_urlsafe(32)
    stamp = now_iso()
    with settings_connection() as db:
        db.execute(
            """
            INSERT INTO employee_sessions (token, employee_id, created_at, last_seen_at)
            VALUES (?, ?, ?, ?)
            """,
            (token, employee_id, stamp, stamp),
        )
    return token


def employee_from_session(payload: dict | None, roles: set[str] | None = None) -> sqlite3.Row | None:
    if not payload:
        return None
    employee_id = parse_int(payload.get("employeeId"), 0)
    token = clean_text(payload.get("sessionToken"))
    if not employee_id or not token:
        return None
    with settings_connection() as db:
        row = db.execute(
            """
            SELECT e.*
            FROM employee_sessions s
            JOIN employees e ON e.id = s.employee_id
            WHERE s.token = ? AND s.employee_id = ? AND e.active = 1
            """,
            (token, employee_id),
        ).fetchone()
        if not row:
            return None
        if roles and row["role"] not in roles:
            return None
        db.execute("UPDATE employee_sessions SET last_seen_at = ? WHERE token = ?", (now_iso(), token))
        return row


def employee_has_permission(row: sqlite3.Row, permission: str) -> bool:
    if permission not in PERMISSION_DEFINITIONS:
        return False
    role = normalize_employee_role(row["role"])
    if role == "admin":
        return True
    return permission in read_role_permissions().get(role, [])


def permission_authorized(payload: dict | None, permission: str) -> bool:
    if valid_admin_password((payload or {}).get("adminPassword")):
        return True
    employee = employee_from_session(payload)
    return bool(employee and employee_has_permission(employee, permission))


def manager_or_admin_authorized(payload: dict | None) -> bool:
    return bool(valid_admin_password((payload or {}).get("adminPassword")) or employee_from_session(payload, {"manager", "admin"}))


def part_report_row(row: sqlite3.Row) -> dict[str, object]:
    payload = part_export_row(row)
    payload["updatedAt"] = row["updated_at"]
    return payload
def get_or_create_brand(db: sqlite3.Connection, name: str) -> int:
    brand_name = clean_text(name)
    if not brand_name:
        raise ValueError("Brand is required.")

    row = db.execute("SELECT id, active, deleted_at FROM brands WHERE name = ?", (brand_name,)).fetchone()
    if row:
        if row["deleted_at"]:
            raise ValueError("That brand was permanently deleted from the app.")
        if not row["active"]:
            raise ValueError("That brand is saved. Restore it from Settings before adding parts to it.")
        return int(row["id"])

    sort_order = db.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM brands").fetchone()[0]
    db.execute(
        "INSERT INTO brands (name, accent, logo, sort_order) VALUES (?, ?, ?, ?)",
        (brand_name, "#2563eb", "", sort_order),
    )
    return int(db.execute("SELECT last_insert_rowid()").fetchone()[0])


def normalize_accent(value: object) -> str:
    accent = clean_text(value)
    if not accent:
        return "#2563eb"
    if not re.fullmatch(r"#[0-9a-fA-F]{6}", accent):
        raise ValueError("Brand color must be a hex color like #2563eb.")
    return accent


def row_to_brand(row: sqlite3.Row) -> dict:
    keys = row.keys()
    return {
        "id": row["id"],
        "name": row["name"],
        "accent": row["accent"],
        "logo": row["logo"],
        "category": row["category"] if "category" in keys else "",
        "defaultFamily": row["default_family"] if "default_family" in keys else "",
        "defaultModel": row["default_model"] if "default_model" in keys else "",
        "defaultCategory": row["default_category"] if "default_category" in keys else "",
        "archiveNote": row["archive_note"] if "archive_note" in keys else "",
        "sortOrder": row["sort_order"],
        "active": bool(row["active"]) if "active" in keys else True,
        "archivedAt": row["archived_at"] if "archived_at" in keys else "",
        "deletedAt": row["deleted_at"] if "deleted_at" in keys else "",
        "deletedName": row["deleted_name"] if "deleted_name" in keys else "",
        "partCount": row["part_count"] or 0,
        "unassignedCount": row["unassigned_count"] or 0,
    }


def build_demo_database(path: Path, department: str) -> None:
    stamp = now_iso()
    db = sqlite3.connect(path)
    try:
        db.executescript(
            """
            CREATE TABLE brands (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                accent TEXT NOT NULL DEFAULT '#2563eb',
                logo TEXT NOT NULL DEFAULT '',
                category TEXT NOT NULL DEFAULT '',
                default_family TEXT NOT NULL DEFAULT '',
                default_model TEXT NOT NULL DEFAULT '',
                default_category TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                archived_at TEXT NOT NULL DEFAULT '',
                archive_note TEXT NOT NULL DEFAULT '',
                deleted_at TEXT NOT NULL DEFAULT '',
                deleted_name TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE parts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
                family TEXT NOT NULL DEFAULT '',
                model TEXT NOT NULL DEFAULT '',
                category TEXT NOT NULL DEFAULT '',
                year_start INTEGER NOT NULL DEFAULT 0,
                year_end INTEGER NOT NULL DEFAULT 0,
                make TEXT NOT NULL DEFAULT '',
                fitment_model TEXT NOT NULL DEFAULT '',
                unit_type TEXT NOT NULL DEFAULT '',
                review_status TEXT NOT NULL DEFAULT 'approved',
                review_note TEXT NOT NULL DEFAULT '',
                item TEXT NOT NULL,
                button_text TEXT NOT NULL DEFAULT '',
                part_number TEXT NOT NULL DEFAULT '',
                old_part_number TEXT NOT NULL DEFAULT '',
                new_part_number TEXT NOT NULL DEFAULT '',
                alternate_numbers TEXT NOT NULL DEFAULT '',
                aftermarket_numbers TEXT NOT NULL DEFAULT '',
                vendor TEXT NOT NULL DEFAULT '',
                tags TEXT NOT NULL DEFAULT '',
                fitment_notes TEXT NOT NULL DEFAULT '',
                attachment_url TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                source TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE service_resources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                resource_type TEXT NOT NULL,
                title TEXT NOT NULL,
                brand TEXT NOT NULL DEFAULT '',
                model TEXT NOT NULL DEFAULT '',
                unit_type TEXT NOT NULL DEFAULT '',
                season TEXT NOT NULL DEFAULT '',
                content TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                username TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'counter',
                pin_hash TEXT NOT NULL DEFAULT '',
                password_hash TEXT NOT NULL DEFAULT '',
                allowed_departments TEXT NOT NULL DEFAULT 'parts,service',
                location_scope TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE UNIQUE INDEX idx_demo_employees_username ON employees(username COLLATE NOCASE) WHERE active = 1 AND username <> '';
            CREATE TABLE employee_sessions (
                token TEXT PRIMARY KEY,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                created_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL
            );
            CREATE TABLE app_settings (key TEXT PRIMARY KEY, value TEXT NOT NULL DEFAULT '', updated_at TEXT NOT NULL);
            """
        )
        brands = [
            ("Can-Am Demo", "#d9232e", "OEM", "Side-by-Side", "Defender", "Maintenance"),
            ("Honda Demo", "#b91c1c", "OEM", "Side-by-Side", "Pioneer", "Maintenance"),
            ("Aftermarket Demo", "#2563eb", "Aftermarket", "Accessories", "", "Filters"),
        ]
        for index, brand in enumerate(brands, start=1):
            db.execute(
                """
                INSERT INTO brands (name, accent, category, default_family, default_model, default_category, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (*brand, index),
            )
        brand_ids = {row[1]: row[0] for row in db.execute("SELECT id, name FROM brands")}
        if department == "parts":
            parts = [
                ("Can-Am Demo", "Maintenance", "Defender", "Filters", 2020, 2026, "Can-Am", "Defender HD10", "UTV", "Oil Filter", "420956123", "OEM", "oil, filter", "Demo common maintenance part"),
                ("Honda Demo", "Maintenance", "Pioneer", "Filters", 2016, 2026, "Honda", "Pioneer 1000", "UTV", "Air Filter", "17254-HL4-A00", "OEM", "air, filter", "Demo fitment part"),
                ("Aftermarket Demo", "Accessories", "Universal", "Battery", 0, 0, "", "", "", "Battery Tender Lead", "BT-081-0069-6", "Aftermarket", "battery, charger", "Demo aftermarket accessory"),
            ]
            for index, part in enumerate(parts, start=1):
                brand, family, model, category, year_start, year_end, make, fitment_model, unit_type, item, number, vendor, tags, notes = part
                db.execute(
                    """
                    INSERT INTO parts (
                        brand_id, family, model, category, year_start, year_end, make, fitment_model, unit_type,
                        item, button_text, part_number, vendor, tags, notes, sort_order, active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (brand_ids[brand], family, model, category, year_start, year_end, make, fitment_model, unit_type, item, item, number, vendor, tags, notes, index, stamp, stamp),
                )
        else:
            resources = [
                ("labor_template", "Oil Change Labor Note", "Can-Am", "Defender HD10", "UTV", "", "Drain oil, replace filter, refill, run, and check for leaks."),
                ("favorite_kit", "Pre-Season Inspection Kit", "", "", "UTV", "Spring", "Oil filter, air filter, spark plugs, belt inspection, tire pressure check."),
                ("model_note", "Pioneer 1000 Service Note", "Honda", "Pioneer 1000", "UTV", "", "Confirm year and submodel before quoting transmission-related parts."),
            ]
            for index, resource in enumerate(resources, start=1):
                db.execute(
                    """
                    INSERT INTO service_resources (resource_type, title, brand, model, unit_type, season, content, sort_order, active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (*resource, index, stamp, stamp),
                )
        demo_employees = [
            ("Demo Admin", "admin", "admin", hash_password("Offroad"), "parts,service", "Training"),
            ("John Doe", "johnd", "counter", hash_password("johnd"), "parts", "Training"),
        ]
        for employee in demo_employees:
            db.execute(
                """
                INSERT INTO employees (name, username, role, password_hash, allowed_departments, location_scope, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (*employee, stamp, stamp),
            )
        db.executemany(
            "INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)",
            [
                ("dealershipName", "Demo Powersports", stamp),
                ("locationName", "Training", stamp),
                ("partsDepartmentLabel", "Parts", stamp),
                ("serviceDepartmentLabel", "Service", stamp),
            ],
        )
        db.commit()
    finally:
        db.close()

def create_demo_database_archive() -> bytes:
    output = io.BytesIO()
    with tempfile.TemporaryDirectory() as directory:
        folder = Path(directory)
        parts_path = folder / "demo-parts.db"
        service_path = folder / "demo-service.db"
        build_demo_database(parts_path, "parts")
        build_demo_database(service_path, "service")
        readme = """Demo Powersports Database

This ZIP contains starter SQLite databases for trying CounterFlow without changing live dealership data.

Files:
- demo-parts.db: sample powersports brands and parts
- demo-service.db: sample service department database

Demo logins in the demo databases:
- admin / Offroad
- johnd / johnd

Use these as examples or keep them as training references.
"""
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.write(parts_path, "demo-parts.db")
            archive.write(service_path, "demo-service.db")
            archive.writestr("README.txt", readme)
    return output.getvalue()


class PartsHandler(BaseHTTPRequestHandler):
    server_version = "CounterFlow/1.0"
    department = DEFAULT_DEPARTMENT

    def db_connection(self) -> sqlite3.Connection:
        return connect(getattr(self, "department", DEFAULT_DEPARTMENT))

    def department_from_request(self, parsed) -> str:
        params = parse_qs(parsed.query)
        return normalize_department(first(params, "department") or self.headers.get("X-PPWork-Department"))


    def do_GET(self) -> None:
        try:
            parsed = urlparse(self.path)
            self.department = self.department_from_request(parsed)
            if parsed.path.startswith("/api/"):
                self.handle_api_get(parsed)
                return
            self.serve_static(parsed.path)
        except Exception as error:
            self.handle_exception(error)

    def do_POST(self) -> None:
        try:
            parsed = urlparse(self.path)
            self.department = self.department_from_request(parsed)
            restore_brand_id = self.restore_brand_id_from_path(parsed.path)
            if restore_brand_id is not None:
                self.restore_brand(restore_brand_id)
                return
            if parsed.path == "/api/setup/admin":
                self.create_first_admin()
                return
            if parsed.path == "/api/employees/login":
                self.login_employee()
                return
            if parsed.path == "/api/employees":
                self.create_employee()
                return
            if parsed.path == "/api/copy-activity":
                self.record_copy_activity_endpoint()
                return
            if parsed.path == "/api/parts":
                self.create_part()
                return
            if parsed.path == "/api/brands":
                self.create_brand()
                return
            if parsed.path == "/api/brands/reorder":
                self.reorder_brands()
                return
            if parsed.path == "/api/admin/backup":
                self.create_backup_endpoint()
                return
            if parsed.path == "/api/admin/restore":
                self.restore_backup_endpoint()
                return
            if parsed.path == "/api/admin/compact":
                self.compact_database_endpoint()
                return
            if parsed.path == "/api/import/parts":
                self.import_parts_csv()
                return
            if parsed.path == "/api/import/parts.xlsx":
                self.import_parts_xlsx()
                return
            if parsed.path == "/api/service-resources":
                self.create_service_resource()
                return
            if parsed.path == "/api/upload-logo":
                self.upload_logo()
                return
            if parsed.path == "/api/reseed":
                self.reseed()
                return
            self.send_error_json(HTTPStatus.NOT_FOUND, "Unknown endpoint.")
        except Exception as error:
            self.handle_exception(error)

    def do_PUT(self) -> None:
        try:
            parsed = urlparse(self.path)
            self.department = self.department_from_request(parsed)
            if parsed.path == "/api/settings/role-permissions":
                self.update_role_permissions()
                return
            if parsed.path == "/api/settings":
                self.update_app_settings()
                return
            if parsed.path == "/api/employee-favorites":
                self.update_employee_favorites()
                return
            employee_id = self.employee_id_from_path(parsed.path)
            if employee_id is not None:
                self.update_employee(employee_id)
                return
            resource_id = self.service_resource_id_from_path(parsed.path)
            if resource_id is not None:
                self.update_service_resource(resource_id)
                return
            brand_id = self.brand_id_from_path(parsed.path)
            if brand_id is not None:
                self.update_brand(brand_id)
                return

            part_id = self.part_id_from_path(parsed.path)
            if part_id is None:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Unknown endpoint.")
                return
            self.update_part(part_id)
        except Exception as error:
            self.handle_exception(error)

    def do_DELETE(self) -> None:
        try:
            parsed = urlparse(self.path)
            self.department = self.department_from_request(parsed)
            employee_id = self.employee_id_from_path(parsed.path)
            if employee_id is not None:
                self.delete_employee(employee_id)
                return
            resource_id = self.service_resource_id_from_path(parsed.path)
            if resource_id is not None:
                self.delete_service_resource(resource_id)
                return
            permanent_brand_id = self.permanent_brand_id_from_path(parsed.path)
            if permanent_brand_id is not None:
                self.permanently_delete_saved_brand(permanent_brand_id)
                return

            brand_id = self.brand_id_from_path(parsed.path)
            if brand_id is not None:
                self.delete_brand(brand_id)
                return

            part_id = self.part_id_from_path(parsed.path)
            if part_id is None:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Unknown endpoint.")
                return
            self.delete_part(part_id)
        except Exception as error:
            self.handle_exception(error)

    def log_message(self, format: str, *args: object) -> None:
        print(f"{self.address_string()} - {format % args}")

    def handle_exception(self, error: BaseException) -> None:
        log_exception(error, method=getattr(self, "command", ""), path=getattr(self, "path", ""), department=getattr(self, "department", ""))
        status, message = friendly_error_message(error)
        try:
            self.send_error_json(status, message)
        except (BrokenPipeError, ConnectionError, OSError):
            pass

    def request_access_payload(self, parsed=None, payload: dict | None = None) -> dict[str, object]:
        access: dict[str, object] = dict(payload or {})
        params = parse_qs(parsed.query) if parsed is not None else {}
        if not access.get("employeeId"):
            access["employeeId"] = first(params, "employeeId") or self.headers.get("X-PPWork-Employee-Id", "")
        if not access.get("sessionToken"):
            access["sessionToken"] = first(params, "sessionToken") or self.headers.get("X-PPWork-Session-Token", "")
        if not access.get("adminPassword"):
            access["adminPassword"] = self.headers.get("X-PPWork-Admin-Password", "")
        return access

    def require_manager_or_admin(self, payload: dict | None, action: str) -> bool:
        if manager_or_admin_authorized(payload):
            return True
        self.send_error_json(HTTPStatus.FORBIDDEN, f"Manager/admin sign-in or admin password is required to {action}.")
        return False

    def require_permission(self, payload: dict | None, permission: str, action: str) -> bool:
        if permission_authorized(payload, permission):
            return True
        label = PERMISSION_DEFINITIONS.get(permission, permission)
        self.send_error_json(HTTPStatus.FORBIDDEN, f"Admin password or {label} role permission is required to {action}.")
        return False

    def handle_api_get(self, parsed) -> None:
        if parsed.path == "/api/departments":
            self.get_departments()
            return
        if parsed.path == "/api/version":
            self.get_version()
            return
        if parsed.path == "/api/release-notes":
            self.get_release_notes()
            return
        if parsed.path == "/api/demo-database":
            self.get_demo_database()
            return
        if parsed.path == "/api/setup/status":
            self.get_setup_status()
            return
        if parsed.path == "/api/settings":
            self.get_app_settings()
            return
        if parsed.path == "/api/deployment-info":
            self.get_deployment_info()
            return
        if parsed.path == "/api/employees":
            self.get_employees()
            return
        if parsed.path == "/api/employee-favorites":
            self.get_employee_favorites(parsed)
            return
        if parsed.path == "/api/service-resources":
            self.get_service_resources(parsed)
            return
        if parsed.path == "/api/reports/review":
            self.get_review_report()
            return
        if parsed.path == "/api/reports/copy-activity":
            self.get_copy_activity_report(parsed)
            return
        if parsed.path == "/api/setup-checklist":
            self.get_setup_checklist()
            return
        if parsed.path == "/api/quick-reference":
            self.get_quick_reference()
            return
        if parsed.path == "/api/network-setup":
            self.get_network_setup()
            return
        if parsed.path == "/api/deployment-checklist":
            self.get_deployment_checklist()
            return
        if parsed.path == "/api/local-link":
            self.get_local_link()
            return
        if parsed.path == "/api/admin/backups":
            self.get_backups()
            return
        if parsed.path == "/api/admin/backup-health":
            self.get_backup_health()
            return
        if parsed.path == "/api/admin/migrations":
            self.get_migrations()
            return
        if parsed.path == "/api/admin/logs":
            self.get_error_log(parsed)
            return
        if parsed.path == "/api/export/parts":
            self.export_parts_csv(parsed)
            return
        if parsed.path == "/api/export/parts.xlsx":
            self.export_parts_xlsx(parsed)
            return
        if parsed.path == "/api/reports/missing":
            self.get_missing_report()
            return
        if parsed.path == "/api/reports/duplicates":
            self.get_duplicate_report()
            return
        if parsed.path == "/api/reports/recent":
            self.get_recent_report(parsed)
            return
        if parsed.path == "/api/reports/print-list":
            self.get_printable_part_list()
            return
        if parsed.path == "/api/brands/saved":
            self.get_saved_brands()
            return
        if parsed.path == "/api/brands":
            self.get_brands()
            return
        if parsed.path == "/api/options":
            self.get_options(parsed)
            return
        if parsed.path == "/api/parts":
            self.get_parts(parsed)
            return
        if parsed.path == "/api/summary":
            self.get_summary()
            return
        self.send_error_json(HTTPStatus.NOT_FOUND, "Unknown endpoint.")

    def get_departments(self) -> None:
        self.send_json([
            {"id": key, "label": value["label"]}
            for key, value in DEPARTMENTS.items()
        ])

    def get_deployment_info(self) -> None:
        self.send_json({
            "appDir": str(BASE_DIR),
            "dataDir": str(DATA_DIR),
            "sharedDataEnabled": DATA_DIR != BASE_DIR,
            "partsDb": str(db_path_for_department("parts")),
            "serviceDb": str(db_path_for_department("service")),
            "backupDir": str(BACKUP_DIR),
            "logFile": str(LOG_FILE),
        })

    def get_version(self) -> None:
        self.send_json({"version": APP_VERSION})

    def get_release_notes(self) -> None:
        path = BASE_DIR / "CHANGELOG.md"
        if not path.exists():
            self.send_error_json(HTTPStatus.NOT_FOUND, "Release notes were not found.")
            return
        self.send_text(path.read_text(encoding="utf-8"), "text/plain; charset=utf-8")

    def get_brands(self) -> None:
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT b.id, b.name, b.accent, b.logo, b.category, b.default_family,
                       b.default_model, b.default_category, b.archive_note, b.sort_order, b.active,
                       b.archived_at, b.deleted_at, b.deleted_name,
                       COUNT(p.id) AS part_count,
                       SUM(CASE WHEN p.active = 1 AND p.part_number = '' THEN 1 ELSE 0 END) AS unassigned_count
                FROM brands b
                LEFT JOIN parts p ON p.brand_id = b.id AND p.active = 1
                WHERE b.active = 1 AND b.deleted_at = ''
                GROUP BY b.id
                ORDER BY b.name COLLATE NOCASE
                """
            ).fetchall()
        self.send_json(
            [row_to_brand(row) for row in rows]
        )

    def get_saved_brands(self) -> None:
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT b.id, b.name, b.accent, b.logo, b.category, b.default_family,
                       b.default_model, b.default_category, b.archive_note, b.sort_order, b.active,
                       b.archived_at, b.deleted_at, b.deleted_name,
                       COUNT(p.id) AS part_count,
                       SUM(CASE WHEN p.active = 1 AND p.part_number = '' THEN 1 ELSE 0 END) AS unassigned_count
                FROM brands b
                LEFT JOIN parts p ON p.brand_id = b.id AND p.active = 1
                WHERE b.active = 0 AND b.deleted_at = ''
                GROUP BY b.id
                ORDER BY b.archived_at DESC, b.name COLLATE NOCASE
                """
            ).fetchall()
        self.send_json([row_to_brand(row) for row in rows])

    def create_brand(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "create brands"):
            return

        try:
            name = clean_text(payload.get("name"))
            if not name:
                raise ValueError("Brand name is required.")
            accent = normalize_accent(payload.get("accent"))
            logo = clean_text(payload.get("logo"))
            category = clean_text(payload.get("category"))
            default_family = clean_text(payload.get("defaultFamily"))
            default_model = clean_text(payload.get("defaultModel"))
            default_category = clean_text(payload.get("defaultCategory"))

            with self.db_connection() as db:
                existing = db.execute("SELECT active, deleted_at FROM brands WHERE name = ?", (name,)).fetchone()
                if existing:
                    if existing["deleted_at"]:
                        self.send_error_json(
                            HTTPStatus.CONFLICT,
                            "A permanently deleted backup record already uses that brand name.",
                        )
                    elif existing["active"]:
                        self.send_error_json(HTTPStatus.CONFLICT, "A brand with that name already exists.")
                    else:
                        self.send_error_json(
                            HTTPStatus.CONFLICT,
                            "A saved brand with that name already exists. Restore it from Saved Brands.",
                        )
                    return
                sort_order = db.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM brands").fetchone()[0]
                cursor = db.execute(
                    """
                    INSERT INTO brands (
                        name, accent, logo, category, default_family, default_model, default_category, sort_order
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (name, accent, logo, category, default_family, default_model, default_category, sort_order),
                )
                brand_id = cursor.lastrowid
        except sqlite3.IntegrityError:
            self.send_error_json(HTTPStatus.CONFLICT, "A brand with that name already exists.")
            return
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"id": brand_id}, HTTPStatus.CREATED)

    def reorder_brands(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "reorder brands"):
            return

        try:
            brand_ids = payload.get("brandIds")
            if not isinstance(brand_ids, list) or not brand_ids:
                raise ValueError("Brand order is required.")
            ordered_ids = [int(brand_id) for brand_id in brand_ids]
            if len(ordered_ids) != len(set(ordered_ids)):
                raise ValueError("Brand order contains duplicate brands.")

            with self.db_connection() as db:
                existing_ids = {
                    int(row["id"])
                    for row in db.execute("SELECT id FROM brands WHERE active = 1 AND deleted_at = ''").fetchall()
                }
                if set(ordered_ids) != existing_ids:
                    raise ValueError("Brand order must include every active brand.")
                for index, brand_id in enumerate(ordered_ids, start=1):
                    db.execute(
                        "UPDATE brands SET sort_order = ? WHERE id = ?",
                        (index, brand_id),
                    )
        except (TypeError, ValueError) as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"ok": True})

    def upload_logo(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "upload brand logos"):
            return

        try:
            data_url = clean_text(payload.get("dataUrl"))
            brand_name = clean_text(payload.get("brandName")) or "brand"
            match = re.fullmatch(r"data:(image/[a-zA-Z0-9.+-]+);base64,(.+)", data_url, re.S)
            if not match:
                raise ValueError("Logo upload must be a valid image file.")

            mime_type = match.group(1).lower()
            extension = LOGO_TYPES.get(mime_type)
            if not extension:
                raise ValueError("Logo must be PNG, JPG, WebP, or GIF.")

            image_bytes = base64.b64decode(match.group(2), validate=True)
            if not image_bytes:
                raise ValueError("Logo file is empty.")
            if len(image_bytes) > MAX_LOGO_BYTES:
                raise ValueError("Logo file must be smaller than 4 MB.")

            logo_bytes, processed_extension, width, height, resized = prepare_logo_image(image_bytes)
            if processed_extension:
                extension = processed_extension
            if len(logo_bytes) > MAX_LOGO_BYTES:
                raise ValueError("Logo file is still too large after processing.")

            slug = re.sub(r"[^a-z0-9]+", "-", brand_name.lower()).strip("-") or "brand"
            stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
            filename = f"{slug}-{stamp}{extension}"
            ASSETS_DIR.mkdir(parents=True, exist_ok=True)
            target = (ASSETS_DIR / filename).resolve()
            if not str(target).startswith(str(ASSETS_DIR.resolve())):
                raise ValueError("Invalid logo filename.")
            target.write_bytes(logo_bytes)
        except (binascii.Error, ValueError) as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"path": f"assets/{filename}", "width": width, "height": height, "resized": resized}, HTTPStatus.CREATED)

    def update_brand(self, brand_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "update brands"):
            return

        try:
            name = clean_text(payload.get("name"))
            if not name:
                raise ValueError("Brand name is required.")
            accent = normalize_accent(payload.get("accent"))
            logo = clean_text(payload.get("logo"))
            category = clean_text(payload.get("category"))
            default_family = clean_text(payload.get("defaultFamily"))
            default_model = clean_text(payload.get("defaultModel"))
            default_category = clean_text(payload.get("defaultCategory"))

            with self.db_connection() as db:
                result = db.execute(
                    """
                    UPDATE brands
                    SET name = ?, accent = ?, logo = ?, category = ?, default_family = ?,
                        default_model = ?, default_category = ?
                    WHERE id = ?
                    """,
                    (name, accent, logo, category, default_family, default_model, default_category, brand_id),
                )
                if result.rowcount == 0:
                    self.send_error_json(HTTPStatus.NOT_FOUND, "Brand not found.")
                    return
        except sqlite3.IntegrityError:
            self.send_error_json(HTTPStatus.CONFLICT, "A brand with that name already exists.")
            return
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"ok": True})

    def delete_brand(self, brand_id: int) -> None:
        payload = self.read_json() if int(self.headers.get("Content-Length") or 0) else {}
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "save and hide brands"):
            return
        archive_note = clean_text(payload.get("archiveNote"))
        backup_database(self.department, "before-brand-delete")
        with self.db_connection() as db:
            row = db.execute(
                """
                SELECT b.name, b.active, COUNT(p.id) AS part_count
                FROM brands b
                LEFT JOIN parts p ON p.brand_id = b.id
                WHERE b.id = ?
                GROUP BY b.id
                """,
                (brand_id,),
            ).fetchone()
            if not row:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Brand not found.")
                return
            if not row["active"]:
                self.send_json({"ok": True})
                return
            db.execute(
                "UPDATE brands SET active = 0, archived_at = ?, archive_note = ? WHERE id = ?",
                (now_iso(), archive_note, brand_id),
            )

        self.send_json({"ok": True})

    def restore_brand(self, brand_id: int) -> None:
        payload = self.read_json() if int(self.headers.get("Content-Length") or 0) else {}
        if payload is None:
            return
        if not self.require_permission(payload, "brandEdit", "restore saved brands"):
            return
        with self.db_connection() as db:
            row = db.execute(
                "SELECT id, active, deleted_at FROM brands WHERE id = ?",
                (brand_id,),
            ).fetchone()
            if not row:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Saved brand not found.")
                return
            if row["deleted_at"]:
                self.send_error_json(HTTPStatus.CONFLICT, "That brand was permanently deleted from the app.")
                return
            if row["active"]:
                self.send_json({"ok": True})
                return

            sort_order = db.execute(
                "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM brands WHERE active = 1 AND deleted_at = ''"
            ).fetchone()[0]
            db.execute(
                """
                UPDATE brands
                SET active = 1, archived_at = '', archive_note = '', sort_order = ?
                WHERE id = ?
                """,
                (sort_order, brand_id),
            )

        self.send_json({"ok": True})

    def permanently_delete_saved_brand(self, brand_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "permanentBrandDelete", "permanently remove saved brands"):
            return

        safety_backup = backup_database(self.department, "before-permanent-brand-removal")
        with self.db_connection() as db:
            row = db.execute(
                """
                SELECT id, name, active, archived_at, deleted_at
                FROM brands
                WHERE id = ?
                """,
                (brand_id,),
            ).fetchone()
            if not row:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Saved brand not found.")
                return
            if row["active"]:
                self.send_error_json(
                    HTTPStatus.CONFLICT,
                    "Save this brand before permanently deleting it.",
                )
                return
            if row["deleted_at"]:
                self.send_json({"ok": True})
                return

            stamp = now_iso()
            backup_name = row["name"]
            hidden_name = f"{backup_name} (deleted {brand_id})"
            db.execute(
                """
                UPDATE brands
                SET name = ?, active = 0,
                    archived_at = CASE WHEN archived_at = '' THEN ? ELSE archived_at END,
                    deleted_at = ?, deleted_name = ?
                WHERE id = ?
                """,
                (hidden_name, stamp, stamp, backup_name, brand_id),
            )

        self.send_json({"ok": True, "backup": safety_backup.name})

    def get_options(self, parsed) -> None:
        params = parse_qs(parsed.query)
        brand = first(params, "brand")

        with self.db_connection() as db:
            args = []
            where = ["p.active = 1", "b.active = 1", "b.deleted_at = ''"]
            if brand:
                where.append("b.name = ?")
                args.append(brand)
            where_sql = " AND ".join(where)

            def values(column: str) -> list[str]:
                rows = db.execute(
                    f"""
                    SELECT DISTINCT p.{column} AS value
                    FROM parts p
                    JOIN brands b ON b.id = p.brand_id
                    WHERE {where_sql} AND p.{column} != ''
                    ORDER BY p.{column} COLLATE NOCASE
                    """,
                    args,
                ).fetchall()
                return [row["value"] for row in rows]

            self.send_json(
                {
                    "families": values("family"),
                    "models": values("model"),
                    "categories": values("category"),
                    "makes": values("make"),
                    "fitmentModels": values("fitment_model"),
                    "unitTypes": values("unit_type"),
                }
            )

    def get_summary(self) -> None:
        with self.db_connection() as db:
            row = db.execute(
                """
                SELECT COUNT(*) AS total,
                       SUM(CASE WHEN p.active = 1 THEN 1 ELSE 0 END) AS active,
                       SUM(CASE WHEN p.active = 1 AND p.part_number = '' THEN 1 ELSE 0 END) AS unassigned
                FROM parts p
                JOIN brands b ON b.id = p.brand_id
                WHERE b.active = 1 AND b.deleted_at = ''
                """
            ).fetchone()
        self.send_json(
            {
                "total": row["total"] or 0,
                "active": row["active"] or 0,
                "unassigned": row["unassigned"] or 0,
            }
        )

    def get_parts(self, parsed) -> None:
        params = parse_qs(parsed.query)
        args = []
        where = ["p.active = 1", "b.active = 1", "b.deleted_at = ''"]

        for key, column in (
            ("brand", "b.name"),
            ("family", "p.family"),
            ("model", "p.model"),
            ("category", "p.category"),
            ("make", "p.make"),
            ("fitmentModel", "p.fitment_model"),
            ("unitType", "p.unit_type"),
        ):
            value = first(params, key)
            if value:
                where.append(f"{column} = ?")
                args.append(value)

        year = parse_int(first(params, "year"), 0)
        if year:
            where.append("(p.year_start = 0 OR p.year_start <= ?)")
            args.append(year)
            where.append("(p.year_end = 0 OR p.year_end >= ?)")
            args.append(year)

        search = first(params, "q")
        with self.db_connection() as db:
            rows = db.execute(
                f"""
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p
                JOIN brands b ON b.id = p.brand_id
                WHERE {" AND ".join(where)}
                ORDER BY b.name COLLATE NOCASE, p.sort_order, p.family, p.model, p.category, p.item
                """,
                args,
            ).fetchall()
        if search:
            rows = [row for row in rows if row_matches_search(row, search)]
        self.send_json([row_to_part(row) for row in rows])

    def create_part(self) -> None:
        payload = self.read_json()
        if payload is None:
            return

        try:
            with self.db_connection() as db:
                brand_id = get_or_create_brand(db, clean_text(payload.get("brand")))
                stamp = now_iso()
                sort_order = db.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM parts WHERE brand_id = ?",
                    (brand_id,),
                ).fetchone()[0]
                db.execute(
                    """
                    INSERT INTO parts (
                        brand_id, family, model, category, year_start, year_end, make,
                        fitment_model, unit_type, review_status, review_note, item, button_text,
                        part_number, old_part_number, new_part_number, alternate_numbers,
                        aftermarket_numbers, vendor, tags, fitment_notes, attachment_url, notes, source,
                        sort_order, active, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?, 1, ?, ?)
                    """,
                    (
                        brand_id,
                        clean_text(payload.get("family")),
                        clean_text(payload.get("model")),
                        clean_text(payload.get("category")),
                        parse_int(payload.get("yearStart"), 0),
                        parse_int(payload.get("yearEnd"), 0),
                        clean_text(payload.get("make")),
                        clean_text(payload.get("fitmentModel")),
                        clean_text(payload.get("unitType")),
                        normalize_review_status(payload.get("reviewStatus")),
                        clean_text(payload.get("reviewNote")),
                        clean_text(payload.get("item")) or "Untitled Part",
                        clean_text(payload.get("buttonText")),
                        clean_text(payload.get("partNumber")),
                        clean_text(payload.get("oldPartNumber")),
                        clean_text(payload.get("newPartNumber")),
                        clean_text(payload.get("alternateNumbers")),
                        clean_text(payload.get("aftermarketNumbers")),
                        clean_text(payload.get("vendor")),
                        clean_text(payload.get("tags")),
                        clean_text(payload.get("fitmentNotes")),
                        clean_text(payload.get("attachmentUrl")),
                        clean_text(payload.get("notes")),
                        sort_order,
                        stamp,
                        stamp,
                    ),
                )
                part_id = int(db.execute("SELECT last_insert_rowid()").fetchone()[0])
                record_part_audit(db, part_id, "create", clean_text(payload.get("item")) or "Untitled Part", after=dict(payload))
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"id": part_id}, HTTPStatus.CREATED)

    def update_part(self, part_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return

        try:
            with self.db_connection() as db:
                existing = db.execute(
                    """
                    SELECT p.*, b.name AS brand, b.accent, b.logo
                    FROM parts p
                    JOIN brands b ON b.id = p.brand_id
                    WHERE p.id = ?
                    """,
                    (part_id,),
                ).fetchone()
                if not existing:
                    self.send_error_json(HTTPStatus.NOT_FOUND, "Part not found.")
                    return

                brand_id = get_or_create_brand(db, clean_text(payload.get("brand")))
                db.execute(
                    """
                    UPDATE parts
                    SET brand_id = ?, family = ?, model = ?, category = ?, year_start = ?, year_end = ?,
                        make = ?, fitment_model = ?, unit_type = ?, review_status = ?, review_note = ?,
                        item = ?, button_text = ?, part_number = ?, old_part_number = ?, new_part_number = ?,
                        alternate_numbers = ?, aftermarket_numbers = ?, vendor = ?, tags = ?, fitment_notes = ?,
                        attachment_url = ?, notes = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        brand_id,
                        clean_text(payload.get("family")),
                        clean_text(payload.get("model")),
                        clean_text(payload.get("category")),
                        parse_int(payload.get("yearStart"), 0),
                        parse_int(payload.get("yearEnd"), 0),
                        clean_text(payload.get("make")),
                        clean_text(payload.get("fitmentModel")),
                        clean_text(payload.get("unitType")),
                        normalize_review_status(payload.get("reviewStatus")),
                        clean_text(payload.get("reviewNote")),
                        clean_text(payload.get("item")) or "Untitled Part",
                        clean_text(payload.get("buttonText")),
                        clean_text(payload.get("partNumber")),
                        clean_text(payload.get("oldPartNumber")),
                        clean_text(payload.get("newPartNumber")),
                        clean_text(payload.get("alternateNumbers")),
                        clean_text(payload.get("aftermarketNumbers")),
                        clean_text(payload.get("vendor")),
                        clean_text(payload.get("tags")),
                        clean_text(payload.get("fitmentNotes")),
                        clean_text(payload.get("attachmentUrl")),
                        clean_text(payload.get("notes")),
                        now_iso(),
                        part_id,
                    ),
                )
                record_part_audit(
                    db,
                    part_id,
                    "update",
                    clean_text(payload.get("item")) or "Untitled Part",
                    before=row_to_part(existing),
                    after=dict(payload),
                )
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return

        self.send_json({"ok": True})

    def delete_part(self, part_id: int) -> None:
        with self.db_connection() as db:
            existing = db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p
                JOIN brands b ON b.id = p.brand_id
                WHERE p.id = ?
                """,
                (part_id,),
            ).fetchone()
            result = db.execute(
                "UPDATE parts SET active = 0, updated_at = ? WHERE id = ?",
                (now_iso(), part_id),
            )
            if result.rowcount == 0:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Part not found.")
                return
            record_part_audit(
                db,
                part_id,
                "delete",
                existing["item"] if existing else "Deleted part",
                before=row_to_part(existing) if existing else {},
            )
        self.send_json({"ok": True})

    def reseed(self) -> None:
        with self.db_connection() as db:
            db.execute("DELETE FROM parts")
            db.execute("DELETE FROM brands")
            db.execute("DELETE FROM sqlite_sequence WHERE name IN ('parts', 'brands')")
            if self.department == DEFAULT_DEPARTMENT:
                seed_database(db)
        self.send_json({"ok": True})

    def get_backups(self) -> None:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        backups = []
        for path in sorted(BACKUP_DIR.glob(f"{self.department}-*.db"), key=lambda item: item.stat().st_mtime, reverse=True):
            backups.append({
                "fileName": path.name,
                "department": self.department,
                "size": path.stat().st_size,
                "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).replace(microsecond=0).isoformat(),
            })
        self.send_json(backups)

    def get_backup_health(self) -> None:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        checked_at = datetime.now(timezone.utc)
        rows = []
        for department in DEPARTMENTS:
            backups = sorted(
                BACKUP_DIR.glob(f"{department}-*.db"),
                key=lambda item: item.stat().st_mtime,
                reverse=True,
            )
            latest = backups[0] if backups else None
            modified_at = ""
            age_hours = None
            size = 0
            status = "Missing"
            if latest is not None:
                modified = datetime.fromtimestamp(latest.stat().st_mtime, timezone.utc).replace(microsecond=0)
                modified_at = modified.isoformat()
                age_hours = round((checked_at - modified).total_seconds() / 3600, 1)
                size = latest.stat().st_size
                if age_hours <= 36:
                    status = "Healthy"
                elif age_hours <= 72:
                    status = "Warning"
                else:
                    status = "Stale"
            rows.append({
                "department": department_label(department),
                "status": status,
                "latestBackup": latest.name if latest else "",
                "ageHours": age_hours,
                "backupCount": len(backups),
                "size": size,
                "modifiedAt": modified_at,
            })
        self.send_json(rows)

    def get_migrations(self) -> None:
        rows = []
        for department in DEPARTMENTS:
            with connect(department) as db:
                for row in db.execute(
                    """
                    SELECT name, app_version, department, applied_at
                    FROM schema_migrations
                    ORDER BY applied_at DESC, id DESC
                    """
                ).fetchall():
                    rows.append({
                        "name": row["name"],
                        "appVersion": row["app_version"],
                        "department": row["department"],
                        "appliedAt": row["applied_at"],
                    })
        rows.sort(key=lambda item: (item["appliedAt"], item["name"]), reverse=True)
        self.send_json(rows)

    def get_error_log(self, parsed) -> None:
        params = parse_qs(parsed.query)
        limit = max(1, min(parse_int(first(params, "limit"), 50), 200))
        if not LOG_FILE.exists():
            self.send_json([])
            return
        lines = LOG_FILE.read_text(encoding="utf-8", errors="replace").splitlines()[-limit:]
        entries = []
        for line in reversed(lines):
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                entries.append({"time": "", "level": "log", "message": line})
        self.send_json(entries)

    def get_demo_database(self) -> None:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        self.send_bytes(
            create_demo_database_archive(),
            "application/zip",
            headers={"Content-Disposition": f'attachment; filename="counterflow-demo-database-{stamp}.zip"'},
        )

    def create_backup_endpoint(self) -> None:
        payload = self.read_json() if int(self.headers.get("Content-Length") or 0) else {}
        if payload is None:
            return
        if not self.require_manager_or_admin(payload, "create backups"):
            return
        target = backup_database(self.department, "manual")
        self.send_json({"ok": True, "fileName": target.name, "path": str(target)})

    def compact_database_endpoint(self) -> None:
        payload = self.read_json() if int(self.headers.get("Content-Length") or 0) else {}
        if payload is None:
            return
        if not self.require_manager_or_admin(payload, "compact and repair databases"):
            return

        db_file = db_path_for_department(self.department)
        safety = backup_database(self.department, "before-compact")
        before_size = db_file.stat().st_size if db_file.exists() else 0
        connection = sqlite3.connect(db_file)
        try:
            integrity = clean_text(connection.execute("PRAGMA integrity_check").fetchone()[0])
            if integrity.lower() != "ok":
                self.send_error_json(
                    HTTPStatus.CONFLICT,
                    f"Database integrity check did not pass. Safety backup created: {safety.name}.",
                )
                return
            connection.execute("PRAGMA optimize")
            connection.commit()
            connection.execute("VACUUM")
        finally:
            connection.close()

        after_size = db_file.stat().st_size if db_file.exists() else 0
        init_db()
        self.send_json({
            "ok": True,
            "department": department_label(self.department),
            "integrity": integrity,
            "beforeSize": before_size,
            "afterSize": after_size,
            "savedBytes": max(before_size - after_size, 0),
            "safetyBackup": safety.name,
        })

    def restore_backup_endpoint(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_manager_or_admin(payload, "restore backups"):
            return
        target = backup_path_from_name(self.department, clean_text(payload.get("fileName")))
        if target is None:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Backup file was not found for this department.")
            return
        safety = backup_database(self.department, "before-restore")
        shutil.copy2(target, db_path_for_department(self.department))
        init_db()
        self.send_json({"ok": True, "restored": target.name, "safetyBackup": safety.name})

    def part_export_rows(self) -> list[sqlite3.Row]:
        with self.db_connection() as db:
            return db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p
                JOIN brands b ON b.id = p.brand_id
                WHERE b.deleted_at = ''
                ORDER BY b.name COLLATE NOCASE, p.active DESC, p.sort_order, p.family, p.model, p.category, p.item
                """
            ).fetchall()

    def export_parts_csv(self, parsed) -> None:
        if not self.require_permission(self.request_access_payload(parsed), "export", "export parts"):
            return
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=PART_EXPORT_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in self.part_export_rows():
            writer.writerow(part_export_row(row))
        filename = f"{self.department}-parts-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
        self.send_text(output.getvalue(), "text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    def export_parts_xlsx(self, parsed) -> None:
        if not self.require_permission(self.request_access_payload(parsed), "export", "export parts"):
            return
        if Workbook is None or Font is None or PatternFill is None:
            self.send_error_json(HTTPStatus.SERVICE_UNAVAILABLE, "Excel export is not available in this Python environment.")
            return
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Parts"
        worksheet.append(PART_EXPORT_COLUMNS)
        worksheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in self.part_export_rows():
            export_row = part_export_row(row)
            worksheet.append([export_row[column] for column in PART_EXPORT_COLUMNS])
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, len(header) + 2, 12), 34)
        output = io.BytesIO()
        workbook.save(output)
        filename = f"{self.department}-parts-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
        self.send_bytes(
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    def import_parts_csv(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "import", "import bulk edits"):
            return
        csv_text = clean_text(payload.get("csvText")).lstrip("\ufeff")
        if not csv_text:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "CSV file is empty.")
            return
        try:
            reader = csv.DictReader(io.StringIO(csv_text))
            if not reader.fieldnames:
                raise ValueError("CSV needs a header row.")
            result = self.import_part_rows(list(reader), "import")
        except (csv.Error, ValueError, sqlite3.Error) as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, f"Import failed: {error}")
            return
        self.send_json(result)

    def import_parts_xlsx(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "import", "import bulk edits"):
            return
        if load_workbook is None:
            self.send_error_json(HTTPStatus.SERVICE_UNAVAILABLE, "Excel import is not available in this Python environment.")
            return
        encoded = clean_text(payload.get("xlsxBase64"))
        if not encoded:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Excel file is empty.")
            return
        try:
            workbook_bytes = base64.b64decode(encoded, validate=True)
            workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
            records = rows_from_worksheet(workbook)
            result = self.import_part_rows(records, "import")
        except (binascii.Error, InvalidFileException, ValueError, sqlite3.Error) as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, f"Excel import failed: {error}")
            return
        self.send_json(result)

    def import_part_rows(self, rows: list[dict[str, str]], backup_reason: str) -> dict[str, object]:
        backup = backup_database(self.department, f"before-{backup_reason}")
        created = updated = skipped = 0
        with self.db_connection() as db:
            for raw in rows:
                payload_row = part_payload_from_csv(raw)
                if not payload_row["brand"] or not payload_row["item"]:
                    skipped += 1
                    continue
                brand_id = get_or_create_brand(db, str(payload_row["brand"]))
                stamp = now_iso()
                part_id = parse_int(csv_value(raw, "id"), 0)
                existing = None
                if part_id:
                    existing = db.execute(
                        """
                        SELECT p.*, b.name AS brand, b.accent, b.logo
                        FROM parts p JOIN brands b ON b.id = p.brand_id
                        WHERE p.id = ?
                        """,
                        (part_id,),
                    ).fetchone()
                if existing:
                    db.execute(
                        """
                        UPDATE parts
                        SET brand_id = ?, family = ?, model = ?, category = ?, year_start = ?, year_end = ?,
                            make = ?, fitment_model = ?, unit_type = ?, review_status = ?, review_note = ?,
                            item = ?, button_text = ?, part_number = ?, old_part_number = ?, new_part_number = ?,
                            alternate_numbers = ?, aftermarket_numbers = ?, vendor = ?, tags = ?, fitment_notes = ?,
                            attachment_url = ?, notes = ?, source = ?, sort_order = ?, active = 1, updated_at = ?
                        WHERE id = ?
                        """,
                        (brand_id, payload_row["family"], payload_row["model"], payload_row["category"],
                         payload_row["yearStart"], payload_row["yearEnd"], payload_row["make"],
                         payload_row["fitmentModel"], payload_row["unitType"], payload_row["reviewStatus"],
                         payload_row["reviewNote"], payload_row["item"], payload_row["buttonText"],
                         payload_row["partNumber"], payload_row["oldPartNumber"], payload_row["newPartNumber"],
                         payload_row["alternateNumbers"], payload_row["aftermarketNumbers"], payload_row["vendor"],
                         payload_row["tags"], payload_row["fitmentNotes"], payload_row["attachmentUrl"],
                         payload_row["notes"], payload_row["source"], payload_row["sortOrder"], stamp, part_id),
                    )
                    record_part_audit(db, part_id, "import-update", str(payload_row["item"]), before=row_to_part(existing), after=payload_row)
                    updated += 1
                else:
                    if not payload_row["sortOrder"]:
                        payload_row["sortOrder"] = db.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM parts WHERE brand_id = ?", (brand_id,)).fetchone()[0]
                    cursor = db.execute(
                        """
                        INSERT INTO parts (
                            brand_id, family, model, category, year_start, year_end, make,
                            fitment_model, unit_type, review_status, review_note, item, button_text,
                            part_number, old_part_number, new_part_number, alternate_numbers,
                            aftermarket_numbers, vendor, tags, fitment_notes, attachment_url, notes, source,
                            sort_order, active, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (brand_id, payload_row["family"], payload_row["model"], payload_row["category"],
                         payload_row["yearStart"], payload_row["yearEnd"], payload_row["make"],
                         payload_row["fitmentModel"], payload_row["unitType"], payload_row["reviewStatus"],
                         payload_row["reviewNote"], payload_row["item"], payload_row["buttonText"],
                         payload_row["partNumber"], payload_row["oldPartNumber"], payload_row["newPartNumber"],
                         payload_row["alternateNumbers"], payload_row["aftermarketNumbers"], payload_row["vendor"],
                         payload_row["tags"], payload_row["fitmentNotes"], payload_row["attachmentUrl"],
                         payload_row["notes"], payload_row["source"], payload_row["sortOrder"], stamp, stamp),
                    )
                    new_id = int(cursor.lastrowid)
                    record_part_audit(db, new_id, "import-create", str(payload_row["item"]), after=payload_row)
                    created += 1
        return {"ok": True, "created": created, "updated": updated, "skipped": skipped, "backup": backup.name}

    def get_setup_status(self) -> None:
        self.send_json({"needsAdminSetup": not admin_employee_exists()})

    def create_first_admin(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if admin_employee_exists():
            self.send_error_json(HTTPStatus.CONFLICT, "First-run setup is already complete.")
            return
        name = clean_text(payload.get("name"))
        username = clean_text(payload.get("username")).lower()
        password = clean_text(payload.get("password"))
        if not name:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Admin name is required.")
            return
        if not username:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Admin username is required.")
            return
        if len(password) < 6:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Admin password must be at least 6 characters.")
            return
        stamp = now_iso()
        try:
            with settings_connection() as db:
                cursor = db.execute(
                    """
                    INSERT INTO employees (
                        name, username, role, pin_hash, password_hash, allowed_departments,
                        location_scope, active, created_at, updated_at
                    ) VALUES (?, ?, 'admin', '', ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        name,
                        username,
                        hash_password(password),
                        normalize_allowed_departments(list(DEPARTMENTS)),
                        "First-run setup",
                        stamp,
                        stamp,
                    ),
                )
                row = db.execute("SELECT * FROM employees WHERE id = ?", (cursor.lastrowid,)).fetchone()
        except sqlite3.IntegrityError:
            self.send_error_json(HTTPStatus.CONFLICT, "An active employee with that name or username already exists.")
            return
        session_token = create_employee_session(int(row["id"]))
        self.send_json(row_to_employee(row, session_token), HTTPStatus.CREATED)

    def get_app_settings(self) -> None:
        self.send_json(app_settings_payload())

    def update_app_settings(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        stamp = now_iso()
        allowed = set(DEFAULT_SETTINGS)
        with settings_connection() as db:
            for key in allowed:
                if key not in payload:
                    continue
                value = role_permissions_json(payload.get(key)) if key == "rolePermissions" else clean_text(payload.get(key))
                db.execute(
                    """
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
                    """,
                    (key, value, stamp),
                )
        self.send_json(app_settings_payload())

    def update_role_permissions(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "employeeEdit", "update role permissions"):
            return
        stamp = now_iso()
        value = role_permissions_json(payload.get("rolePermissions"))
        with settings_connection() as db:
            db.execute(
                """
                INSERT INTO app_settings (key, value, updated_at)
                VALUES ('rolePermissions', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
                """,
                (value, stamp),
            )
        self.send_json(app_settings_payload())

    def get_employees(self) -> None:
        with settings_connection() as db:
            rows = db.execute(
                """
                SELECT * FROM employees
                WHERE active = 1
                ORDER BY name COLLATE NOCASE
                """
            ).fetchall()
        self.send_json([row_to_employee(row) for row in rows])

    def create_employee(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "employeeEdit", "save employees"):
            return
        name = clean_text(payload.get("name"))
        if not name:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Employee name is required.")
            return
        username = clean_text(payload.get("username")).lower()
        role = normalize_employee_role(payload.get("role"))
        stamp = now_iso()
        try:
            with settings_connection() as db:
                cursor = db.execute(
                    """
                    INSERT INTO employees (
                        name, username, role, pin_hash, password_hash, allowed_departments,
                        location_scope, active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        name,
                        username,
                        role,
                        hash_pin(payload.get("pin")),
                        hash_password(payload.get("password")),
                        normalize_allowed_departments(payload.get("allowedDepartments")),
                        clean_text(payload.get("locationScope")),
                        stamp,
                        stamp,
                    ),
                )
                row = db.execute("SELECT * FROM employees WHERE id = ?", (cursor.lastrowid,)).fetchone()
        except sqlite3.IntegrityError:
            self.send_error_json(HTTPStatus.CONFLICT, "An active employee with that name or username already exists.")
            return
        self.send_json(row_to_employee(row), HTTPStatus.CREATED)

    def update_employee(self, employee_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "employeeEdit", "save employees"):
            return
        name = clean_text(payload.get("name"))
        if not name:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Employee name is required.")
            return
        username = clean_text(payload.get("username")).lower()
        role = normalize_employee_role(payload.get("role"))
        stamp = now_iso()
        try:
            with settings_connection() as db:
                existing = db.execute("SELECT * FROM employees WHERE id = ? AND active = 1", (employee_id,)).fetchone()
                if not existing:
                    self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
                    return
                pin_hash = existing["pin_hash"]
                if clean_text(payload.get("pin")):
                    pin_hash = hash_pin(payload.get("pin"))
                if payload.get("clearPin"):
                    pin_hash = ""
                password_hash = existing["password_hash"] if "password_hash" in existing.keys() else ""
                if clean_text(payload.get("password")):
                    password_hash = hash_password(payload.get("password"))
                if payload.get("clearPassword"):
                    password_hash = ""
                db.execute(
                    """
                    UPDATE employees
                    SET name = ?, username = ?, role = ?, pin_hash = ?, password_hash = ?,
                        allowed_departments = ?, location_scope = ?, updated_at = ?
                    WHERE id = ? AND active = 1
                    """,
                    (
                        name,
                        username,
                        role,
                        pin_hash,
                        password_hash,
                        normalize_allowed_departments(payload.get("allowedDepartments")),
                        clean_text(payload.get("locationScope")),
                        stamp,
                        employee_id,
                    ),
                )
                row = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        except sqlite3.IntegrityError:
            self.send_error_json(HTTPStatus.CONFLICT, "An active employee with that name or username already exists.")
            return
        self.send_json(row_to_employee(row))

    def delete_employee(self, employee_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return
        if not self.require_permission(payload, "employeeEdit", "delete employees"):
            return
        with settings_connection() as db:
            cursor = db.execute(
                "UPDATE employees SET active = 0, updated_at = ? WHERE id = ? AND active = 1",
                (now_iso(), employee_id),
            )
        if cursor.rowcount == 0:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
            return
        self.send_json({"ok": True})

    def login_employee(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        employee_id = parse_int(payload.get("employeeId"), 0)
        username = clean_text(payload.get("username")).lower()
        name = clean_text(payload.get("name"))
        with settings_connection() as db:
            if username:
                row = db.execute("SELECT * FROM employees WHERE username = ? COLLATE NOCASE AND active = 1", (username,)).fetchone()
            elif employee_id:
                row = db.execute("SELECT * FROM employees WHERE id = ? AND active = 1", (employee_id,)).fetchone()
            else:
                row = db.execute("SELECT * FROM employees WHERE name = ? COLLATE NOCASE AND active = 1", (name,)).fetchone()
        if not row:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
            return
        password_hash = row["password_hash"] if "password_hash" in row.keys() else ""
        if password_hash:
            if not verify_password(payload.get("password"), password_hash):
                self.send_error_json(HTTPStatus.FORBIDDEN, "Employee password did not match.")
                return
        elif not verify_pin(payload.get("pin"), row["pin_hash"]):
            self.send_error_json(HTTPStatus.FORBIDDEN, "Employee PIN did not match.")
            return
        session_token = create_employee_session(int(row["id"]))
        self.send_json(row_to_employee(row, session_token))

    def get_employee_favorites(self, parsed) -> None:
        params = parse_qs(parsed.query)
        employee_id = parse_int(first(params, "employeeId"), 0)
        with settings_connection() as db:
            employee = db.execute("SELECT id FROM employees WHERE id = ? AND active = 1", (employee_id,)).fetchone()
            if not employee:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
                return
            rows = db.execute(
                """
                SELECT part_id FROM employee_favorites
                WHERE employee_id = ? AND department = ?
                ORDER BY created_at
                """,
                (employee_id, self.department),
            ).fetchall()
        self.send_json({"partIds": [str(row["part_id"]) for row in rows]})

    def update_employee_favorites(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        employee_id = parse_int(payload.get("employeeId"), 0)
        raw_ids = payload.get("partIds")
        if not isinstance(raw_ids, list):
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Favorite part IDs are required.")
            return
        part_ids: list[int] = []
        for value in raw_ids:
            part_id = parse_int(value, 0)
            if part_id > 0 and part_id not in part_ids:
                part_ids.append(part_id)
        stamp = now_iso()
        with settings_connection() as db:
            employee = db.execute("SELECT id FROM employees WHERE id = ? AND active = 1", (employee_id,)).fetchone()
            if not employee:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
                return
            db.execute(
                "DELETE FROM employee_favorites WHERE employee_id = ? AND department = ?",
                (employee_id, self.department),
            )
            db.executemany(
                """
                INSERT INTO employee_favorites (employee_id, department, part_id, created_at)
                VALUES (?, ?, ?, ?)
                """,
                [(employee_id, self.department, part_id, stamp) for part_id in part_ids],
            )
        self.send_json({"ok": True, "count": len(part_ids)})

    def record_copy_activity_endpoint(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        employee_id = parse_int(payload.get("employeeId"), 0)
        part_id = parse_int(payload.get("partId"), 0)
        with settings_connection() as settings_db:
            employee = settings_db.execute("SELECT * FROM employees WHERE id = ? AND active = 1", (employee_id,)).fetchone()
        if not employee:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Employee not found.")
            return
        with self.db_connection() as db:
            part = db.execute(
                """
                SELECT p.id, p.item, p.part_number, b.name AS brand
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE p.id = ? AND p.active = 1
                """,
                (part_id,),
            ).fetchone()
        if not part:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Part not found.")
            return
        with settings_connection() as settings_db:
            settings_db.execute(
                """
                INSERT INTO employee_copy_activity (
                    employee_id, employee_name, department, part_id, brand, item, part_number, copied_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    employee_id, employee["name"], self.department, part_id,
                    part["brand"], part["item"], part["part_number"], now_iso(),
                ),
            )
        self.send_json({"ok": True})

    def get_copy_activity_report(self, parsed) -> None:
        params = parse_qs(parsed.query)
        limit = max(1, min(parse_int(first(params, "limit"), 100), 500))
        with settings_connection() as db:
            rows = db.execute(
                """
                SELECT employee_id, employee_name, department, part_id, brand, item, part_number, copied_at
                FROM employee_copy_activity
                WHERE department = ?
                ORDER BY copied_at DESC
                LIMIT ?
                """,
                (self.department, limit),
            ).fetchall()
        self.send_json([
            {
                "employeeId": row["employee_id"],
                "employeeName": row["employee_name"],
                "department": row["department"],
                "partId": row["part_id"],
                "brand": row["brand"],
                "item": row["item"],
                "partNumber": row["part_number"],
                "copiedAt": row["copied_at"],
            }
            for row in rows
        ])

    def get_local_link(self) -> None:
        host = clean_text(self.headers.get("Host")) or "127.0.0.1:8765"
        local_url = f"http://{host}/"
        network_urls = []
        try:
            port = host.split(":")[-1] if ":" in host else "8765"
            for address in socket.gethostbyname_ex(socket.gethostname())[2]:
                if address and not address.startswith("127."):
                    network_urls.append(f"http://{address}:{port}/")
        except OSError:
            network_urls = []
        self.send_json({"localUrl": local_url, "networkUrls": sorted(set(network_urls))})

    def get_setup_checklist(self) -> None:
        settings = read_app_settings()
        content = f"""# Powersports Parts Board Setup Checklist

Dealership: {settings.get('dealershipName') or 'Independence County Offroad'}
Location: {settings.get('locationName') or 'Main'}

## Counter Setup

- Start the app on the main counter computer.
- Confirm Parts and Service departments open from Settings.
- Open the local link from every counter computer that needs access.
- Add employee accounts for counter, manager, and admin users.
- Choose each employee's theme, density, pinned brands, and favorites on their workstation.

## Catalog Setup

- Add active OEM and aftermarket brands.
- Upload brand logos where helpful.
- Export the Parts catalog to Excel.
- Fill in missing part numbers and powersports fitment fields.
- Import the edited Excel file with the admin password.
- Run Missing Numbers, Duplicates, Review Queue, and Recent Changes reports.

## Service Setup

- Add labor note templates for common jobs.
- Add Service favorite kits for oil, belt, brake, tire, battery, and inspection work.
- Add model-specific service notes for common ATVs, UTVs, and motorcycles.
- Add seasonal service packages for pre-season, winterization, and post-ride inspections.

## Safety

- Create a manual backup after initial setup.
- Keep the admin password limited to managers.
- Review the Copy Activity report when signed-in employees need follow-up.
- Use the review queue for uncertain part numbers before employees copy them at the counter.
"""
        self.send_text(content, "text/markdown; charset=utf-8")

    def get_quick_reference(self) -> None:
        settings = read_app_settings()
        dealership = html.escape(settings.get("dealershipName") or "Independence County Offroad")
        parts_label = html.escape(settings.get("partsDepartmentLabel") or "Parts")
        service_label = html.escape(settings.get("serviceDepartmentLabel") or "Service")
        document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><title>{dealership} Quick Reference</title>
<style>body{{font-family:Arial,sans-serif;margin:24px;color:#111827}}h1{{margin-bottom:4px}}h2{{margin-top:22px}}table{{border-collapse:collapse;width:100%;font-size:13px}}th,td{{border:1px solid #d1d5db;padding:8px;text-align:left;vertical-align:top}}th{{background:#f3f4f6}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}li{{margin:6px 0}}@media print{{body{{margin:10mm}}button{{display:none}}.grid{{grid-template-columns:1fr}}}}</style>
</head><body><button onclick="window.print()">Print</button><h1>{dealership} Parts Board</h1><p>{parts_label} and {service_label} quick reference</p><div class="grid"><section><h2>Counter Workflow</h2><ol><li>Choose the correct department.</li><li>Search by part, model, part number, vendor, tag, or fitment.</li><li>Click a part card to copy its part number.</li><li>Paste the copied number into the CRM, cashier, or repair order screen.</li><li>Use the star next to a part number for employee favorites.</li></ol></section><section><h2>Review Workflow</h2><ol><li>Mark uncertain parts as Needs Review.</li><li>Open Settings and run Review Queue.</li><li>Update the part number or note after confirmation.</li><li>Set the part back to Approved.</li></ol></section></div><h2>Useful Buttons</h2><table><thead><tr><th>Area</th><th>Use</th></tr></thead><tbody><tr><td>Favorites</td><td>Personal saved parts for the signed-in employee.</td></tr><tr><td>Recently Copied</td><td>Fast access to numbers copied on this workstation.</td></tr><tr><td>Most Used</td><td>Commonly copied parts on this workstation.</td></tr><tr><td>Missing Numbers</td><td>Parts that still need a current part number.</td></tr><tr><td>Copy Activity</td><td>Signed-in employee copy history for manager follow-up.</td></tr></tbody></table></body></html>"""
        self.send_text(document, "text/html; charset=utf-8")

    def get_network_setup(self) -> None:
        settings = read_app_settings()
        dealership = html.escape(settings.get("dealershipName") or "Independence County Offroad")
        host = clean_text(self.headers.get("Host")) or "127.0.0.1:8765"
        port = host.split(":")[-1] if ":" in host else "8765"
        links = [f"http://{html.escape(host)}/"]
        try:
            for address in socket.gethostbyname_ex(socket.gethostname())[2]:
                if address and not address.startswith("127."):
                    links.append(f"http://{html.escape(address)}:{html.escape(port)}/")
        except OSError:
            pass
        link_rows = "".join(f"<li><code>{link}</code></li>" for link in sorted(set(links)))
        document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><title>{dealership} Network Setup</title>
<style>body{{font-family:Arial,sans-serif;margin:24px;color:#111827;line-height:1.45}}h1{{margin-bottom:4px}}code{{background:#f3f4f6;padding:2px 5px;border-radius:4px}}li{{margin:7px 0}}@media print{{body{{margin:10mm}}button{{display:none}}}}</style>
</head><body><button onclick="window.print()">Print</button><h1>{dealership} CounterFlow Network Setup</h1><p>Use these steps for counter computers that open this app from the main computer.</p><h2>Main Computer</h2><ol><li>Start CounterFlow on the main counter computer.</li><li>Leave the app window running during business hours.</li><li>Use a wired connection or reliable dealership Wi-Fi where possible.</li></ol><h2>Counter Computers</h2><ol><li>Open one of these links in the browser:</li></ol><ul>{link_rows}</ul><h2>Checklist</h2><ol><li>Confirm Parts and Service both open.</li><li>Sign in each employee, then choose their theme and density.</li><li>Copy one known part number and paste it into the CRM test screen.</li><li>Create a manual backup after setup is complete.</li></ol></body></html>"""
        self.send_text(document, "text/html; charset=utf-8")

    def get_deployment_checklist(self) -> None:
        settings = read_app_settings()
        dealership = html.escape(settings.get("dealershipName") or "Independence County Offroad")
        data_dir = html.escape(str(DATA_DIR))
        app_dir = html.escape(str(BASE_DIR))
        backup_dir = html.escape(str(BACKUP_DIR))
        log_file = html.escape(str(LOG_FILE))
        parts_db = html.escape(str(db_path_for_department("parts")))
        service_db = html.escape(str(db_path_for_department("service")))
        document = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><title>{dealership} Deployment Checklist</title>
<style>body{{font-family:Arial,sans-serif;margin:24px;color:#111827;line-height:1.45}}h1{{margin-bottom:4px}}h2{{margin-top:22px}}code{{background:#f3f4f6;padding:2px 5px;border-radius:4px}}li{{margin:7px 0}}table{{border-collapse:collapse;width:100%;font-size:13px}}th,td{{border:1px solid #d1d5db;padding:8px;text-align:left;vertical-align:top}}th{{background:#f3f4f6}}@media print{{body{{margin:10mm}}button{{display:none}}}}</style>
</head><body><button onclick="window.print()">Print</button><h1>{dealership} CounterFlow Deployment Checklist</h1><p>Use this before putting CounterFlow on a shared counter computer or moving it to a server.</p><h2>Current Paths</h2><table><tbody><tr><th>App Folder</th><td><code>{app_dir}</code></td></tr><tr><th>Data Folder</th><td><code>{data_dir}</code></td></tr><tr><th>Parts Database</th><td><code>{parts_db}</code></td></tr><tr><th>Service Database</th><td><code>{service_db}</code></td></tr><tr><th>Backups</th><td><code>{backup_dir}</code></td></tr><tr><th>Log File</th><td><code>{log_file}</code></td></tr></tbody></table><h2>Shared Folder Setup</h2><ol><li>Place the app folder on the computer that will stay on during counter hours.</li><li>Confirm the data folder is backed up by OneDrive, Windows backup, or a dealership backup tool.</li><li>Start the app and create a manual backup from Settings.</li><li>Open the Network Setup sheet and test the link on each counter workstation.</li></ol><h2>Single Host Network Setup</h2><ol><li>Use the most reliable parts-counter computer or a small office server as the host.</li><li>Keep the host awake during business hours.</li><li>Allow the app port through Windows Firewall only for the dealership network.</li><li>Create employee logins before handing out the network link.</li></ol><h2>Cloud Hosted Setup</h2><ol><li>Move the data folder to persistent storage, not temporary server storage.</li><li>Use HTTPS and a real admin password before exposing the app outside the local network.</li><li>Schedule daily database backups and test a restore before launch.</li><li>Restrict employee access to dealership users only.</li></ol><h2>Go Live Checks</h2><ol><li>Run Backup Health and confirm each department has a current backup.</li><li>Export the Parts catalog to Excel and store a copy with launch records.</li><li>Sign in as a counter employee and copy one known part number into the CRM test screen.</li><li>Confirm each workstation has its own theme and favorites behavior.</li></ol></body></html>"""
        self.send_text(document, "text/html; charset=utf-8")

    def get_service_resources(self, parsed) -> None:
        params = parse_qs(parsed.query)
        resource_type = clean_text(first(params, "type"))
        args = []
        where = ["active = 1"]
        if resource_type:
            where.append("resource_type = ?")
            args.append(normalize_service_resource_type(resource_type))
        with self.db_connection() as db:
            rows = db.execute(
                f"""
                SELECT * FROM service_resources
                WHERE {' AND '.join(where)}
                ORDER BY sort_order, title COLLATE NOCASE
                """,
                args,
            ).fetchall()
        self.send_json([row_to_service_resource(row) for row in rows])

    def create_service_resource(self) -> None:
        payload = self.read_json()
        if payload is None:
            return
        try:
            resource_type = normalize_service_resource_type(payload.get("type"))
            title = clean_text(payload.get("title"))
            content = clean_text(payload.get("content"))
            if not title:
                raise ValueError("Service workflow title is required.")
            stamp = now_iso()
            with self.db_connection() as db:
                sort_order = db.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM service_resources WHERE resource_type = ?",
                    (resource_type,),
                ).fetchone()[0]
                cursor = db.execute(
                    """
                    INSERT INTO service_resources (
                        resource_type, title, brand, model, unit_type, season, content,
                        sort_order, active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        resource_type, title, clean_text(payload.get("brand")), clean_text(payload.get("model")),
                        clean_text(payload.get("unitType")), clean_text(payload.get("season")), content,
                        sort_order, stamp, stamp,
                    ),
                )
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return
        self.send_json({"id": int(cursor.lastrowid)}, HTTPStatus.CREATED)

    def update_service_resource(self, resource_id: int) -> None:
        payload = self.read_json()
        if payload is None:
            return
        try:
            resource_type = normalize_service_resource_type(payload.get("type"))
            title = clean_text(payload.get("title"))
            if not title:
                raise ValueError("Service workflow title is required.")
            with self.db_connection() as db:
                cursor = db.execute(
                    """
                    UPDATE service_resources
                    SET resource_type = ?, title = ?, brand = ?, model = ?, unit_type = ?,
                        season = ?, content = ?, updated_at = ?
                    WHERE id = ? AND active = 1
                    """,
                    (
                        resource_type, title, clean_text(payload.get("brand")), clean_text(payload.get("model")),
                        clean_text(payload.get("unitType")), clean_text(payload.get("season")),
                        clean_text(payload.get("content")), now_iso(), resource_id,
                    ),
                )
            if cursor.rowcount == 0:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Service workflow item not found.")
                return
        except ValueError as error:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(error))
            return
        self.send_json({"ok": True})

    def delete_service_resource(self, resource_id: int) -> None:
        with self.db_connection() as db:
            cursor = db.execute(
                "UPDATE service_resources SET active = 0, updated_at = ? WHERE id = ?",
                (now_iso(), resource_id),
            )
        if cursor.rowcount == 0:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Service workflow item not found.")
            return
        self.send_json({"ok": True})

    def get_review_report(self) -> None:
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE p.active = 1 AND b.active = 1 AND b.deleted_at = '' AND p.review_status = 'needs-review'
                ORDER BY b.name COLLATE NOCASE, p.updated_at DESC, p.item
                """
            ).fetchall()
        self.send_json([part_report_row(row) for row in rows])

    def get_missing_report(self) -> None:
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE p.active = 1 AND b.active = 1 AND b.deleted_at = '' AND p.part_number = ''
                ORDER BY b.name COLLATE NOCASE, p.family, p.model, p.category, p.item
                """
            ).fetchall()
        self.send_json([part_report_row(row) for row in rows])

    def get_duplicate_report(self) -> None:
        duplicates = []
        with self.db_connection() as db:
            numbers = db.execute(
                """
                SELECT p.part_number, COUNT(*) AS part_count
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE p.active = 1 AND b.active = 1 AND b.deleted_at = '' AND p.part_number != ''
                GROUP BY p.part_number HAVING COUNT(*) > 1
                ORDER BY part_count DESC, p.part_number
                """
            ).fetchall()
            for number in numbers:
                parts = db.execute(
                    """
                    SELECT p.*, b.name AS brand, b.accent, b.logo
                    FROM parts p JOIN brands b ON b.id = p.brand_id
                    WHERE p.active = 1 AND b.active = 1 AND b.deleted_at = '' AND p.part_number = ?
                    ORDER BY b.name COLLATE NOCASE, p.family, p.model, p.category, p.item
                    """,
                    (number["part_number"],),
                ).fetchall()
                duplicates.append({"partNumber": number["part_number"], "count": number["part_count"], "parts": [part_report_row(row) for row in parts]})
        self.send_json(duplicates)

    def get_recent_report(self, parsed) -> None:
        params = parse_qs(parsed.query)
        limit = max(1, min(parse_int(first(params, "limit"), 50), 200))
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE b.deleted_at = ''
                ORDER BY p.updated_at DESC, p.id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        self.send_json([part_report_row(row) for row in rows])

    def get_printable_part_list(self) -> None:
        with self.db_connection() as db:
            rows = db.execute(
                """
                SELECT p.*, b.name AS brand, b.accent, b.logo
                FROM parts p JOIN brands b ON b.id = p.brand_id
                WHERE p.active = 1 AND b.active = 1 AND b.deleted_at = ''
                ORDER BY b.name COLLATE NOCASE, p.family, p.model, p.category, p.item
                """
            ).fetchall()
        title = f"{department_label(self.department)} Parts List"
        table_rows = []
        for row in rows:
            table_rows.append(
                "<tr>"
                f"<td>{html.escape(row['brand'])}</td>"
                f"<td>{html.escape(row['family'])}</td>"
                f"<td>{html.escape(row['model'])}</td>"
                f"<td>{html.escape(row['category'])}</td>"
                f"<td>{html.escape(row['item'])}</td>"
                f"<td>{html.escape(row['part_number'] or 'Needs number')}</td>"
                f"<td>{html.escape(row['vendor'])}</td>"
                "</tr>"
            )
        document = f"""<!doctype html><html lang=\"en\"><head><meta charset=\"utf-8\"><title>{html.escape(title)}</title>
<style>body{{font-family:Arial,sans-serif;margin:24px;color:#111827}}table{{border-collapse:collapse;width:100%;font-size:12px}}th,td{{border:1px solid #d1d5db;padding:6px 8px;text-align:left;vertical-align:top}}th{{background:#f3f4f6}}@media print{{body{{margin:10mm}}button{{display:none}}}}</style>
</head><body><button onclick=\"window.print()\">Print</button><h1>{html.escape(title)}</h1><p>{len(rows)} active parts</p><table><thead><tr><th>Brand</th><th>Family</th><th>Model</th><th>Category</th><th>Item</th><th>Part Number</th><th>Vendor</th></tr></thead><tbody>{''.join(table_rows)}</tbody></table></body></html>"""
        self.send_text(document, "text/html; charset=utf-8")

    def read_json(self) -> dict | None:
        length = int(self.headers.get("Content-Length") or 0)
        if length > MAX_JSON_BYTES:
            self.send_error_json(HTTPStatus.REQUEST_ENTITY_TOO_LARGE, "Request is too large.")
            return None
        try:
            raw = self.rfile.read(length).decode("utf-8")
            return json.loads(raw or "{}")
        except json.JSONDecodeError:
            self.send_error_json(HTTPStatus.BAD_REQUEST, "Invalid JSON.")
            return None

    def service_resource_id_from_path(self, path: str) -> int | None:
        match = re.fullmatch(r"/api/service-resources/(\d+)/?", path)
        if not match:
            return None
        return int(match.group(1))

    def employee_id_from_path(self, path: str) -> int | None:
        match = re.fullmatch(r"/api/employees/(\d+)/?", path)
        if not match:
            return None
        return int(match.group(1))

    def part_id_from_path(self, path: str) -> int | None:
        prefix = "/api/parts/"
        if not path.startswith(prefix):
            return None
        try:
            return int(path.removeprefix(prefix).strip("/"))
        except ValueError:
            return None

    def brand_id_from_path(self, path: str) -> int | None:
        prefix = "/api/brands/"
        if not path.startswith(prefix):
            return None
        try:
            return int(path.removeprefix(prefix).strip("/"))
        except ValueError:
            return None

    def restore_brand_id_from_path(self, path: str) -> int | None:
        match = re.fullmatch(r"/api/brands/(\d+)/restore/?", path)
        if not match:
            return None
        return int(match.group(1))

    def permanent_brand_id_from_path(self, path: str) -> int | None:
        match = re.fullmatch(r"/api/brands/(\d+)/permanent/?", path)
        if not match:
            return None
        return int(match.group(1))

    def serve_static(self, url_path: str) -> None:
        if url_path in {"", "/"}:
            relative = "index.html"
        else:
            relative = unquote(url_path.lstrip("/"))

        target = (STATIC_DIR / relative).resolve()
        if not str(target).startswith(str(STATIC_DIR.resolve())):
            self.send_error(HTTPStatus.FORBIDDEN)
            return

        if not target.exists() or not target.is_file():
            if relative.startswith("assets/"):
                log_event("warning", "Missing asset requested", path=relative)
                self.send_text("Asset not found.", "text/plain; charset=utf-8", HTTPStatus.NOT_FOUND)
                return
            target = STATIC_DIR / "index.html"

        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        data = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload: object, status: HTTPStatus = HTTPStatus.OK) -> None:
        data = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_text(
        self,
        content: str,
        content_type: str,
        status: HTTPStatus = HTTPStatus.OK,
        headers: dict[str, str] | None = None,
    ) -> None:
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def send_bytes(
        self,
        data: bytes,
        content_type: str,
        status: HTTPStatus = HTTPStatus.OK,
        headers: dict[str, str] | None = None,
    ) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def send_error_json(self, status: HTTPStatus, message: str) -> None:
        self.send_json({"error": message}, status)


def first(params: dict[str, list[str]], key: str) -> str:
    values = params.get(key, [])
    if not values:
        return ""
    return clean_text(values[0])


def env_int(name: str, fallback: int) -> int:
    value = os.environ.get(name)
    if not value:
        return fallback
    try:
        return int(value)
    except ValueError:
        return fallback
def main() -> None:
    default_host = os.environ.get("PPWORK_HOST", "127.0.0.1")
    default_port = env_int("PORT", env_int("PPWORK_PORT", 8765))
    parser = argparse.ArgumentParser(description="Run the CounterFlow parts and service counter board.")
    parser.add_argument("--host", default=default_host)
    parser.add_argument("--port", default=default_port, type=int)
    args = parser.parse_args()

    init_db()
    create_daily_startup_backups()
    server = ThreadingHTTPServer((args.host, args.port), PartsHandler)
    print(f"CounterFlow is running at http://{args.host}:{args.port}")
    print(f"Data folder: {DATA_DIR}")
    for value in DEPARTMENTS.values():
        print(f"{value['label']} database: {value['db']}")
    server.serve_forever()


if __name__ == "__main__":
    main()
